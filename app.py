import os
from datetime import datetime, date, timedelta

from flask import (
    Flask, render_template, redirect, url_for,
    request, flash, jsonify, abort, send_file, current_app
)
from flask_login import (
    LoginManager, login_user, logout_user,
    login_required, current_user
)

from sqlalchemy import or_
from flask_wtf.csrf import CSRFProtect, generate_csrf
from flask_cors import CORS

from config import Config
from models import db, User, Branch, Store, Category, Item
from models import Customer, Supplier, Purchase, PurchaseItem
from models import Sale, SaleItem, Return, ReturnItem
from models import Expense, Bond, StoreSetting, Plan, UserSubscription, Employee, ActivityLog, PasswordResetToken, Notification, UserSetting
from email_utils import send_verification_email, send_password_reset_email
import secrets

DROPDOWN_LIMIT = 500

app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)
csrf = CSRFProtect(app)
CORS(app, resources={r"/api/*": {"origins": "*"}})

from api_routes import api, verify_token
csrf.exempt(api)
app.register_blueprint(api)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'يرجى تسجيل الدخول أولاً'
login_manager.login_message_category = 'warning'


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


def generate_invoice(prefix):
    today = date.today().strftime('%Y%m%d')
    count = 1
    while True:
        inv = f'{prefix}-{today}-{count:04d}'
        if prefix == 'SALE':
            if not Sale.query.filter_by(invoice_number=inv).first():
                return inv
        elif prefix == 'PUR':
            if not Purchase.query.filter_by(invoice_number=inv).first():
                return inv
        elif prefix == 'RET':
            if not Return.query.filter_by(return_number=inv).first():
                return inv
        elif prefix == 'BND':
            if not Bond.query.filter_by(bond_number=inv).first():
                return inv
        count += 1


def get_dashboard_stats(user_id=None):
    base_filter = {} if user_id is None else {'user_id': user_id}

    total_sales = db.session.query(db.func.sum(Sale.total)).filter_by(**base_filter).scalar() or 0
    total_purchases = db.session.query(db.func.sum(Purchase.total)).filter_by(**base_filter).scalar() or 0
    total_expenses = db.session.query(db.func.sum(Expense.amount)).filter_by(**base_filter).scalar() or 0
    total_bonds_receipt = db.session.query(db.func.sum(Bond.amount)).filter(Bond.bond_type == 'receipt', Bond.user_id == user_id if user_id else True).scalar() or 0
    total_bonds_payment = db.session.query(db.func.sum(Bond.amount)).filter(Bond.bond_type == 'payment', Bond.user_id == user_id if user_id else True).scalar() or 0
    total_returns = db.session.query(db.func.sum(Return.total)).filter_by(**base_filter).scalar() or 0

    cogs_filter = [SaleItem.sale_id.in_(
        db.session.query(Sale.id).filter_by(**base_filter)
    )] if user_id else []
    cogs = db.session.query(
        db.func.sum(SaleItem.quantity * db.func.coalesce(Item.buy_price, 0))
    ).join(Item, SaleItem.item_id == Item.id, isouter=True
    ).filter(*cogs_filter).scalar() or 0

    return {
        'customers_count': Customer.query.filter_by(is_active=True, **base_filter).count(),
        'suppliers_count': Supplier.query.filter_by(is_active=True, **base_filter).count(),
        'branches_count': Branch.query.filter_by(is_active=True, **base_filter).count(),
        'stores_count': Store.query.filter_by(is_active=True, **base_filter).count(),
        'items_count': Item.query.filter_by(is_active=True, **base_filter).count(),
        'total_sales': total_sales,
        'total_purchases': total_purchases,
        'total_expenses': total_expenses,
        'total_bonds_receipt': total_bonds_receipt,
        'total_bonds_payment': total_bonds_payment,
        'total_returns': total_returns,
        'sales_count': Sale.query.filter_by(**base_filter).count(),
        'purchases_count': Purchase.query.filter_by(**base_filter).count(),
        'returns_count': Return.query.filter_by(**base_filter).count(),
        'expenses_count': Expense.query.filter_by(**base_filter).count(),
        'bonds_receipt_count': Bond.query.filter_by(bond_type='receipt', **base_filter).count(),
        'bonds_payment_count': Bond.query.filter_by(bond_type='payment', **base_filter).count(),
        'cogs': round(cogs, 2),
        'gross_profit': round(total_sales - cogs, 2),
        'profit': round(total_sales - cogs - total_expenses, 2),
        'net_revenue': round(total_sales - total_returns, 2),
        'recent_sales': Sale.query.filter_by(**base_filter).order_by(Sale.created_at.desc()).limit(5).all(),
        'recent_purchases': Purchase.query.filter_by(**base_filter).order_by(Purchase.created_at.desc()).limit(5).all(),
    }


def get_month_range(months_ago):
    """Get start and end datetime for a month N months ago from now."""
    now = datetime.utcnow()
    total_months = now.year * 12 + (now.month - 1) - months_ago
    year = total_months // 12
    month = (total_months % 12) + 1
    month_start = datetime(year, month, 1)
    if months_ago == 0:
        month_end = now
    else:
        next_total = total_months + 1
        ny = next_total // 12
        nm = (next_total % 12) + 1
        month_end = datetime(ny, nm, 1)
    return month_start, month_end


def get_monthly_analytics(user_id=None):
    """Return monthly sales, purchases, expenses, and profit data for last 6 months."""
    base = {} if user_id is None else {'user_id': user_id}
    months = []
    for i in range(5, -1, -1):
        month_start, month_end = get_month_range(i)
        m_label = month_start.strftime('%b')
        month_sales = db.session.query(db.func.sum(Sale.total)).filter(
            Sale.created_at >= month_start, Sale.created_at < month_end
        ).filter_by(**base).scalar() or 0
        month_purchases = db.session.query(db.func.sum(Purchase.total)).filter(
            Purchase.created_at >= month_start, Purchase.created_at < month_end
        ).filter_by(**base).scalar() or 0
        month_expenses = db.session.query(db.func.sum(Expense.amount)).filter(
            Expense.created_at >= month_start, Expense.created_at < month_end
        ).filter_by(**base).scalar() or 0
        sale_ids_q = db.session.query(Sale.id).filter(
            Sale.created_at >= month_start, Sale.created_at < month_end
        ).filter_by(**base)
        month_cogs = db.session.query(
            db.func.sum(SaleItem.quantity * db.func.coalesce(Item.buy_price, 0))
        ).join(Item, SaleItem.item_id == Item.id, isouter=True
        ).filter(SaleItem.sale_id.in_(sale_ids_q)).scalar() or 0
        month_profit = round(month_sales - month_cogs - month_expenses, 2)
        months.append({
            'label': m_label,
            'sales': round(month_sales, 2),
            'purchases': round(month_purchases, 2),
            'expenses': round(month_expenses, 2),
            'cogs': round(month_cogs, 2),
            'profit': month_profit,
        })
    return months


def get_top_items(limit=5, user_id=None):
    """Get top selling items by quantity."""
    from sqlalchemy import desc
    base = {} if user_id is None else {'user_id': user_id}
    results = db.session.query(
        Item.name,
        db.func.sum(SaleItem.quantity).label('total_qty'),
        db.func.sum(SaleItem.total).label('total_revenue'),
    ).join(SaleItem, SaleItem.item_id == Item.id
    ).join(Sale, Sale.id == SaleItem.sale_id
    ).filter(Sale.id.in_(
        db.session.query(Sale.id).filter_by(**base)
    ) if user_id else True
    ).group_by(Item.id, Item.name
    ).order_by(desc('total_qty')
    ).limit(limit).all()
    return [{'name': r[0], 'quantity': r[1], 'revenue': round(r[2], 2)} for r in results]


def get_chart_data(user_id=None):
    """Get daily sales data for last 30 days for chart rendering."""
    from datetime import timedelta
    base = {} if user_id is None else {'user_id': user_id}
    now = datetime.utcnow()
    data = []
    for i in range(14, -1, -1):
        day = (now - timedelta(days=i)).date()
        day_start = datetime.combine(day, datetime.min.time())
        day_end = datetime.combine(day, datetime.max.time())
        daily_total = db.session.query(db.func.sum(Sale.total)).filter(
            Sale.created_at >= day_start, Sale.created_at <= day_end
        ).filter_by(**base).scalar() or 0
        data.append({
            'day': day.strftime('%d %b'),
            'total': round(daily_total, 2),
        })
    return data


def get_stock_alerts(user_id=None):
    """Get items with low or out of stock."""
    base = {} if user_id is None else {'user_id': user_id}
    low_stock = Item.query.filter_by(is_active=True, **base).filter(
        Item.quantity > 0,
        Item.quantity <= Item.min_quantity
    ).count()
    out_of_stock = Item.query.filter_by(is_active=True, **base).filter(
        Item.quantity <= 0
    ).count()
    return {'low_stock': low_stock, 'out_of_stock': out_of_stock}


def get_recent_returns(limit=5, user_id=None):
    base = {} if user_id is None else {'user_id': user_id}
    return Return.query.filter_by(**base).order_by(Return.created_at.desc()).limit(limit).all()


def get_top_customers(limit=5, user_id=None):
    from sqlalchemy import desc
    base = {} if user_id is None else {'user_id': user_id}
    results = db.session.query(
        Customer.name,
        db.func.count(Sale.id).label('total_invoices'),
        db.func.sum(Sale.total).label('total_amount'),
        db.func.sum(Sale.paid).label('total_paid'),
    ).join(Sale, Sale.customer_id == Customer.id
    ).filter(Sale.id.in_(
        db.session.query(Sale.id).filter_by(**base)
    ) if user_id else True
    ).group_by(Customer.id, Customer.name
    ).order_by(desc('total_amount')
    ).limit(limit).all()
    return [{
        'name': r[0],
        'invoices': r[1],
        'total': round(r[2], 2),
        'paid': round(r[3] or 0, 2),
        'balance': round(r[2] - (r[3] or 0), 2),
    } for r in results]


def get_returns_monthly(user_id=None):
    base = {} if user_id is None else {'user_id': user_id}
    months = []
    for i in range(5, -1, -1):
        month_start, month_end = get_month_range(i)
        m_label = month_start.strftime('%b')
        month_returns = db.session.query(db.func.sum(Return.total)).filter(
            Return.created_at >= month_start, Return.created_at < month_end
        ).filter_by(**base).scalar() or 0
        month_count = Return.query.filter_by(**base).filter(
            Return.created_at >= month_start, Return.created_at < month_end
        ).count()
        months.append({
            'label': m_label,
            'total': round(month_returns, 2),
            'count': month_count,
        })
    return months


def get_sales_monthly(user_id=None):
    base = {} if user_id is None else {'user_id': user_id}
    months = []
    for i in range(5, -1, -1):
        month_start, month_end = get_month_range(i)
        m_label = month_start.strftime('%b')
        month_sales = db.session.query(db.func.sum(Sale.total)).filter(
            Sale.created_at >= month_start, Sale.created_at < month_end
        ).filter_by(**base).scalar() or 0
        month_count = Sale.query.filter_by(**base).filter(
            Sale.created_at >= month_start, Sale.created_at < month_end
        ).count()
        month_paid = db.session.query(db.func.sum(Sale.paid)).filter(
            Sale.created_at >= month_start, Sale.created_at < month_end
        ).filter_by(**base).scalar() or 0
        months.append({
            'label': m_label,
            'total': round(month_sales, 2),
            'count': month_count,
            'paid': round(month_paid, 2),
            'remaining': round(month_sales - month_paid, 2),
        })
    return months


def get_purchases_monthly(user_id=None):
    base = {} if user_id is None else {'user_id': user_id}
    months = []
    for i in range(5, -1, -1):
        month_start, month_end = get_month_range(i)
        m_label = month_start.strftime('%b')
        month_purchases = db.session.query(db.func.sum(Purchase.total)).filter(
            Purchase.created_at >= month_start, Purchase.created_at < month_end
        ).filter_by(**base).scalar() or 0
        month_count = Purchase.query.filter_by(**base).filter(
            Purchase.created_at >= month_start, Purchase.created_at < month_end
        ).count()
        month_paid = db.session.query(db.func.sum(Purchase.paid)).filter(
            Purchase.created_at >= month_start, Purchase.created_at < month_end
        ).filter_by(**base).scalar() or 0
        months.append({
            'label': m_label,
            'total': round(month_purchases, 2),
            'count': month_count,
            'paid': round(month_paid, 2),
            'remaining': round(month_purchases - month_paid, 2),
        })
    return months


# ─────────────────────────── Error Handlers ───────────────────────────

@app.errorhandler(404)
def not_found(error):
    return render_template('landing.html'), 404



# ─────────────────────────── Auth Routes ───────────────────────────

@app.route('/')
def landing():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('landing.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        if not email or not password:
            flash('يرجى إدخال البريد الإلكتروني وكلمة المرور', 'danger')
            return render_template('login.html')
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            if not user.is_active:
                flash('هذا الحساب غير نشط', 'danger')
                return render_template('login.html')
            login_user(user, remember=request.form.get('remember'))
            user.last_login = datetime.utcnow()
            safe_commit()
            flash(f'مرحباً {user.full_name}', 'success')
            # للموظفين: وجههم لأول صفحة متاحة
            if user.role == 'employee':
                emp = Employee.query.filter_by(user_id=user.id, is_active=True).first()
                if emp:
                    target = find_first_available_page(emp)
                    if target:
                        return redirect(url_for(target))
            return redirect(url_for('dashboard'))
        flash('البريد الإلكتروني أو كلمة المرور غير صحيحة', 'danger')
    return render_template('login.html')


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '').strip().lower()
        phone = request.form.get('phone', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        if not full_name or not email or not password:
            flash('جميع الحقول المطلوبة يجب أن تمتلئ', 'danger')
            return render_template('signup.html')
        if password != confirm_password:
            flash('كلمة المرور غير متطابقة', 'danger')
            return render_template('signup.html')
        if len(password) < 6:
            flash('كلمة المرور يجب أن تكون 6 أحرف على الأقل', 'danger')
            return render_template('signup.html')
        if User.query.filter_by(email=email).first():
            flash('البريد الإلكتروني مستخدم بالفعل', 'danger')
            return render_template('signup.html')
        user = User(full_name=full_name, email=email, phone=phone, role='admin')
        user.set_password(password)
        db.session.add(user)
        db.session.flush()
        # Create default data for new user
        default_branch = Branch(name='الفرع الرئيسي', address='', phone='', manager='', user_id=user.id)
        db.session.add(default_branch)
        db.session.flush()
        default_store = Store(name='المخزن الرئيسي', branch_id=default_branch.id, address='', manager='', user_id=user.id)
        db.session.add(default_store)
        default_category = Category(name='عام', description='تصنيف عام', user_id=user.id)
        db.session.add(default_category)
        settings = StoreSetting(user_id=user.id, store_name='متجري', store_phone=phone, store_address='', store_email=email, currency='ج.م', invoice_footer='شكراً لتسوقكم')
        db.session.add(settings)
        trial_plan = Plan.query.filter_by(is_active=True).order_by(Plan.price).first()
        if trial_plan and trial_plan.trial_days and trial_plan.trial_days > 0:
            trial_sub = UserSubscription(
                user_id=user.id,
                plan_id=trial_plan.id,
                start_date=datetime.utcnow(),
                end_date=datetime.utcnow() + timedelta(days=trial_plan.trial_days),
                is_active=True,
                payment_status='trial',
                payment_method='auto',
                notes=f'اشتراك تجريبي لمدة {trial_plan.trial_days} يوم',
            )
            db.session.add(trial_sub)
        # Send verification email
        token = secrets.token_urlsafe(32)
        user.verification_token = token
        safe_commit()
        send_verification_email(user, token)
        flash(f'تم إنشاء الحساب بنجاح! تم تفعيل اشتراك تجريبي لمدة {trial_plan.trial_days if trial_plan else 7} يوم. يرجى التحقق من بريدك الإلكتروني لتأكيد الحساب', 'success')
        return redirect(url_for('login'))
    return render_template('signup.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('تم تسجيل الخروج بنجاح', 'info')
    return redirect(url_for('landing'))


@app.route('/verify-email/<token>')
def verify_email(token):
    user = User.query.filter_by(verification_token=token).first()
    if not user:
        flash('رابط التأكيد غير صالح أو منتهي الصلاحية', 'danger')
        return redirect(url_for('login'))
    user.email_verified = True
    user.verification_token = None
    safe_commit()
    flash('تم تأكيد البريد الإلكتروني بنجاح! يمكنك تسجيل الدخول الآن', 'success')
    return redirect(url_for('login'))


@app.route('/auth/google')
def auth_google():
    from authlib.integrations.flask_client import OAuth
    google_client_id = current_app.config.get('GOOGLE_CLIENT_ID', '')
    google_client_secret = current_app.config.get('GOOGLE_CLIENT_SECRET', '')
    if not google_client_id or not google_client_secret:
        flash('تسجيل الدخول عبر Google غير مفعل حالياً', 'warning')
        return redirect(url_for('login'))
    redirect_uri = url_for('auth_google_callback', _external=True)
    oauth = OAuth(current_app)
    oauth.register(
        name='google',
        client_id=google_client_id,
        client_secret=google_client_secret,
        access_token_url='https://accounts.google.com/o/oauth2/token',
        access_token_params=None,
        authorize_url='https://accounts.google.com/o/oauth2/auth',
        authorize_params=None,
        api_base_url='https://www.googleapis.com/oauth2/v1/',
        userinfo_endpoint='https://www.googleapis.com/oauth2/v3/userinfo',
        client_kwargs={'scope': 'openid email profile'},
    )
    return oauth.google.authorize_redirect(redirect_uri)


@app.route('/auth/google/callback')
def auth_google_callback():
    from authlib.integrations.flask_client import OAuth
    google_client_id = current_app.config.get('GOOGLE_CLIENT_ID', '')
    google_client_secret = current_app.config.get('GOOGLE_CLIENT_SECRET', '')
    if not google_client_id or not google_client_secret:
        flash('تسجيل الدخول عبر Google غير مفعل حالياً', 'warning')
        return redirect(url_for('login'))
    oauth = OAuth(current_app)
    oauth.register(
        name='google',
        client_id=google_client_id,
        client_secret=google_client_secret,
        access_token_url='https://accounts.google.com/o/oauth2/token',
        access_token_params=None,
        authorize_url='https://accounts.google.com/o/oauth2/auth',
        authorize_params=None,
        api_base_url='https://www.googleapis.com/oauth2/v1/',
        userinfo_endpoint='https://www.googleapis.com/oauth2/v3/userinfo',
        client_kwargs={'scope': 'openid email profile'},
    )
    try:
        token = oauth.google.authorize_access_token()
        userinfo = oauth.google.get('https://www.googleapis.com/oauth2/v3/userinfo').json()
        google_id = userinfo.get('sub')
        email = userinfo.get('email', '').lower()
        name = userinfo.get('name', '')
        if not google_id or not email:
            flash('فشل الحصول على بيانات Google', 'danger')
            return redirect(url_for('login'))
        user = User.query.filter_by(google_id=google_id).first()
        if not user:
            user = User.query.filter_by(email=email).first()
            if user:
                user.google_id = google_id
            else:
                user = User(
                    full_name=name or email.split('@')[0],
                    email=email,
                    phone='',
                    password_hash='',
                    role='admin',
                    google_id=google_id,
                    email_verified=True,
                )
                user.set_password(secrets.token_urlsafe(16))
                db.session.add(user)
                db.session.flush()
                default_branch = Branch(name='الفرع الرئيسي', address='', phone='', manager='', user_id=user.id)
                db.session.add(default_branch)
                db.session.flush()
                default_store = Store(name='المخزن الرئيسي', branch_id=default_branch.id, address='', manager='', user_id=user.id)
                db.session.add(default_store)
                default_category = Category(name='عام', description='تصنيف عام', user_id=user.id)
                db.session.add(default_category)
                settings = StoreSetting(user_id=user.id, store_name='متجري', store_phone='', store_address='', store_email=email, currency='ج.م', invoice_footer='شكراً لتسوقكم')
                db.session.add(settings)
                trial_plan = Plan.query.filter_by(is_active=True).order_by(Plan.price).first()
                if trial_plan and trial_plan.trial_days and trial_plan.trial_days > 0:
                    trial_sub = UserSubscription(
                        user_id=user.id, plan_id=trial_plan.id,
                        start_date=datetime.utcnow(),
                        end_date=datetime.utcnow() + timedelta(days=trial_plan.trial_days),
                        is_active=True, payment_status='trial', payment_method='auto',
                    )
                    db.session.add(trial_sub)
        db.session.commit()
        login_user(user)
        flash(f'مرحباً {user.full_name}', 'success')
        return redirect(url_for('dashboard'))
    except Exception as e:
        current_app.logger.error(f'Google auth error: {e}')
        flash('فشل تسجيل الدخول عبر Google', 'danger')
        return redirect(url_for('login'))


# ─────────────────────────── Password Reset ───────────────────────────

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        user = User.query.filter_by(email=email).first()
        if user:
            token = PasswordResetToken(user_id=user.id)
            db.session.add(token)
            safe_commit()
            send_password_reset_email(user, token.token)
        flash('إذا كان البريد الإلكتروني مسجلاً، ستتلقى رابط إعادة التعيين', 'info')
        return redirect(url_for('login'))
    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    reset_token = PasswordResetToken.query.filter_by(token=token).first()
    if not reset_token or not reset_token.is_valid():
        flash('رابط إعادة تعيين كلمة المرور غير صالح أو منتهي الصلاحية', 'danger')
        return redirect(url_for('forgot_password'))
    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        if not password or len(password) < 6:
            flash('كلمة المرور يجب أن تكون 6 أحرف على الأقل', 'danger')
            return render_template('reset_password.html', token=token)
        if password != confirm:
            flash('كلمة المرور غير متطابقة', 'danger')
            return render_template('reset_password.html', token=token)
        user = reset_token.user
        user.set_password(password)
        reset_token.used = True
        safe_commit()
        flash('تم تغيير كلمة المرور بنجاح. يمكنك تسجيل الدخول الآن', 'success')
        return redirect(url_for('login'))
    return render_template('reset_password.html', token=token)


# ─────────────────────────── Activity Logs ───────────────────────────

@app.route('/activity-logs')
@login_required
def activity_logs():
    page = request.args.get('page', 1, type=int)
    query = ActivityLog.query
    if not current_user.is_super_admin:
        query = query.filter_by(user_id=current_user.id)
    pagination = query.order_by(ActivityLog.created_at.desc()).paginate(page=page, per_page=PER_PAGE, error_out=False)
    return render_template('activity_logs.html', user=current_user, logs=pagination.items, pagination=pagination)


# ─────────────────────────── Backup Database ───────────────────────────

@app.route('/backup')
@login_required
def backup_database():
    if current_user.role == 'employee':
        flash('صلاحية النسخ الاحتياطي غير متاحة للموظفين', 'danger')
        return redirect(url_for('dashboard'))
    db_uri = current_app.config['SQLALCHEMY_DATABASE_URI']
    if db_uri.startswith('sqlite:///'):
        db_path = db_uri[len('sqlite:///'):]
    else:
        db_path = os.path.join(current_app.instance_path, 'mazad_plus.db')
    if not os.path.exists(db_path):
        flash('ملف قاعدة البيانات غير موجود', 'danger')
        return redirect(url_for('dashboard'))
    today_str = date.today().strftime('%Y-%m-%d')
    filename = f'mazadplus_backup_{today_str}.db'
    return send_file(db_path, as_attachment=True, download_name=filename)


# ─────────────────────────── Dashboard ───────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    uid = current_user.id if not current_user.is_super_admin else None
    stats = get_dashboard_stats(uid)
    monthly = get_monthly_analytics(uid)
    top_items = get_top_items(5, uid)
    chart_data = get_chart_data(uid)
    alerts = get_stock_alerts(uid)
    recent_returns = get_recent_returns(5, uid)
    top_customers = get_top_customers(5, uid)
    returns_monthly = get_returns_monthly(uid)
    sales_monthly = get_sales_monthly(uid)
    purchases_monthly = get_purchases_monthly(uid)
    return render_template('dashboard.html', user=current_user, stats=stats,
                           monthly=monthly, top_items=top_items,
                           chart_data=chart_data, alerts=alerts,
                           recent_returns=recent_returns, top_customers=top_customers,
                           returns_monthly=returns_monthly, sales_monthly=sales_monthly,
                           purchases_monthly=purchases_monthly)


# ═══════════════════════════════════════════════
# BRANCHES CRUD
# ═══════════════════════════════════════════════

PER_PAGE = 20

@app.route('/branches')
@login_required
def branches_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()
    query = Branch.query.filter_by(user_id=current_user.id)
    if search:
        query = query.filter(Branch.name.ilike(f'%{search}%'))
    pagination = query.order_by(Branch.name).paginate(page=page, per_page=PER_PAGE, error_out=False)
    return render_template('branches.html', user=current_user, branches=pagination.items, pagination=pagination)


@app.route('/branches/add', methods=['POST'])
@login_required
def branch_add():
    if not check_plan_limit('branches'):
        flash('لقد تجاوزت الحد المسموح به من الفروع في باقتك', 'danger')
        return redirect(url_for('branches_list'))
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم الفرع مطلوب', 'danger')
        return redirect(url_for('branches_list'))
    branch = Branch(
        name=name,
        address=request.form.get('address', '').strip(),
        phone=request.form.get('phone', '').strip(),
        manager=request.form.get('manager', '').strip(),
        user_id=current_user.id,
    )
    db.session.add(branch)
    log_activity('create', 'فرع', branch.id, name)
    safe_commit()
    flash(f'تم إضافة الفرع {name} بنجاح', 'success')
    return redirect(url_for('branches_list'))


@app.route('/branches/edit/<int:id>', methods=['POST'])
@login_required
def branch_edit(id):
    branch = Branch.query.get_or_404(id)
    if branch.user_id and branch.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم الفرع مطلوب', 'danger')
        return redirect(url_for('branches_list'))
    branch.name = name
    branch.address = request.form.get('address', '').strip()
    branch.phone = request.form.get('phone', '').strip()
    branch.manager = request.form.get('manager', '').strip()
    log_activity('update', 'فرع', branch.id, name)
    safe_commit()
    flash(f'تم تحديث الفرع {name} بنجاح', 'success')
    return redirect(url_for('branches_list'))


@app.route('/branches/delete/<int:id>', methods=['POST'])
@login_required
def branch_delete(id):
    branch = Branch.query.get_or_404(id)
    if branch.user_id and branch.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    name = branch.name
    db.session.delete(branch)
    log_activity('delete', 'فرع', id, name)
    safe_commit()
    flash(f'تم حذف الفرع {name} بنجاح', 'success')
    return redirect(url_for('branches_list'))


# ═══════════════════════════════════════════════
# STORES (المخازن) CRUD
# ═══════════════════════════════════════════════

@app.route('/stores')
@login_required
def stores_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()
    query = Store.query.filter_by(user_id=current_user.id)
    if search:
        query = query.filter(Store.name.ilike(f'%{search}%'))
    pagination = query.order_by(Store.name).paginate(page=page, per_page=PER_PAGE, error_out=False)
    branches = Branch.query.filter_by(is_active=True, user_id=current_user.id).all()
    return render_template('stores.html', user=current_user, stores=pagination.items, branches=branches, pagination=pagination)


@app.route('/stores/add', methods=['POST'])
@login_required
def store_add():
    if not check_plan_limit('stores'):
        flash('لقد تجاوزت الحد المسموح به من المخازن في باقتك', 'danger')
        return redirect(url_for('stores_list'))
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم المخزن مطلوب', 'danger')
        return redirect(url_for('stores_list'))
    store = Store(
        name=name,
        branch_id=request.form.get('branch_id', type=int),
        address=request.form.get('address', '').strip(),
        manager=request.form.get('manager', '').strip(),
        user_id=current_user.id,
    )
    db.session.add(store)
    log_activity('create', 'مخزن', store.id, name)
    safe_commit()
    flash(f'تم إضافة المخزن {name} بنجاح', 'success')
    return redirect(url_for('stores_list'))


@app.route('/stores/edit/<int:id>', methods=['POST'])
@login_required
def store_edit(id):
    store = Store.query.get_or_404(id)
    if store.user_id and store.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم المخزن مطلوب', 'danger')
        return redirect(url_for('stores_list'))
    store.name = name
    store.branch_id = request.form.get('branch_id', type=int)
    store.address = request.form.get('address', '').strip()
    store.manager = request.form.get('manager', '').strip()
    log_activity('update', 'مخزن', store.id, name)
    safe_commit()
    flash(f'تم تحديث المخزن {name} بنجاح', 'success')
    return redirect(url_for('stores_list'))


@app.route('/stores/delete/<int:id>', methods=['POST'])
@login_required
def store_delete(id):
    store = Store.query.get_or_404(id)
    if store.user_id and store.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    name = store.name
    db.session.delete(store)
    log_activity('delete', 'مخزن', id, name)
    safe_commit()
    flash(f'تم حذف المخزن {name} بنجاح', 'success')
    return redirect(url_for('stores_list'))


# ═══════════════════════════════════════════════
# CATEGORIES (الأصناف - تصنيفات) CRUD
# ═══════════════════════════════════════════════

@app.route('/categories')
@login_required
def categories_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()
    query = Category.query.filter_by(user_id=current_user.id)
    if search:
        query = query.filter(Category.name.ilike(f'%{search}%'))
    pagination = query.order_by(Category.name).paginate(page=page, per_page=PER_PAGE, error_out=False)
    return render_template('categories.html', user=current_user, categories=pagination.items, pagination=pagination)


@app.route('/categories/add', methods=['POST'])
@login_required
def category_add():
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم التصنيف مطلوب', 'danger')
        return redirect(url_for('categories_list'))
    category = Category(name=name, description=request.form.get('description', '').strip(), user_id=current_user.id)
    db.session.add(category)
    log_activity('create', 'تصنيف', category.id, name)
    safe_commit()
    flash(f'تم إضافة التصنيف {name} بنجاح', 'success')
    return redirect(url_for('categories_list'))


@app.route('/categories/edit/<int:id>', methods=['POST'])
@login_required
def category_edit(id):
    category = Category.query.get_or_404(id)
    if category.user_id and category.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم التصنيف مطلوب', 'danger')
        return redirect(url_for('categories_list'))
    category.name = name
    category.description = request.form.get('description', '').strip()
    log_activity('update', 'تصنيف', category.id, name)
    safe_commit()
    flash(f'تم تحديث التصنيف {name} بنجاح', 'success')
    return redirect(url_for('categories_list'))


@app.route('/categories/delete/<int:id>', methods=['POST'])
@login_required
def category_delete(id):
    category = Category.query.get_or_404(id)
    if category.user_id and category.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    name = category.name
    db.session.delete(category)
    log_activity('delete', 'تصنيف', id, name)
    safe_commit()
    flash(f'تم حذف التصنيف {name} بنجاح', 'success')
    return redirect(url_for('categories_list'))


# ═══════════════════════════════════════════════
# ITEMS (الأصناف) CRUD
# ═══════════════════════════════════════════════

@app.route('/items')
@login_required
def items_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()
    query = Item.query.filter_by(user_id=current_user.id)
    if search:
        query = query.filter(
            or_(Item.name.ilike(f'%{search}%'), Item.barcode.ilike(f'%{search}%'))
        )
    pagination = query.order_by(Item.name).paginate(page=page, per_page=PER_PAGE, error_out=False)
    categories = Category.query.filter_by(is_active=True, user_id=current_user.id).all()
    stores = Store.query.filter_by(is_active=True, user_id=current_user.id).all()
    return render_template('items.html', user=current_user, items=pagination.items,
                           categories=categories, stores=stores, pagination=pagination)


@app.route('/items/add', methods=['POST'])
@login_required
def item_add():
    if not check_plan_limit('items'):
        flash('لقد تجاوزت الحد المسموح به من الأصناف في باقتك', 'danger')
        return redirect(url_for('items_list'))
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم الصنف مطلوب', 'danger')
        return redirect(url_for('items_list'))
    item = Item(
        name=name,
        barcode=request.form.get('barcode', '').strip(),
        category_id=request.form.get('category_id', type=int),
        store_id=request.form.get('store_id', type=int),
        buy_price=request.form.get('buy_price', 0, type=float),
        sell_price=request.form.get('sell_price', 0, type=float),
        quantity=request.form.get('quantity', 0, type=int),
        min_quantity=request.form.get('min_quantity', 0, type=int),
        description=request.form.get('description', '').strip(),
        user_id=current_user.id,
    )
    db.session.add(item)
    log_activity('create', 'صنف', item.id, name)
    safe_commit()
    flash(f'تم إضافة الصنف {name} بنجاح', 'success')
    return redirect(url_for('items_list'))


@app.route('/items/edit/<int:id>', methods=['POST'])
@login_required
def item_edit(id):
    item = Item.query.get_or_404(id)
    if item.user_id and item.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم الصنف مطلوب', 'danger')
        return redirect(url_for('items_list'))
    item.name = name
    item.barcode = request.form.get('barcode', '').strip()
    item.category_id = request.form.get('category_id', type=int)
    item.store_id = request.form.get('store_id', type=int)
    item.buy_price = request.form.get('buy_price', 0, type=float)
    item.sell_price = request.form.get('sell_price', 0, type=float)
    item.quantity = request.form.get('quantity', 0, type=int)
    item.min_quantity = request.form.get('min_quantity', 0, type=int)
    item.description = request.form.get('description', '').strip()
    log_activity('update', 'صنف', item.id, name)
    safe_commit()
    flash(f'تم تحديث الصنف {name} بنجاح', 'success')
    return redirect(url_for('items_list'))


@app.route('/items/delete/<int:id>', methods=['POST'])
@login_required
def item_delete(id):
    item = Item.query.get_or_404(id)
    if item.user_id and item.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    name = item.name
    db.session.delete(item)
    log_activity('delete', 'صنف', id, name)
    safe_commit()
    flash(f'تم حذف الصنف {name} بنجاح', 'success')
    return redirect(url_for('items_list'))


# ═══════════════════════════════════════════════
# CUSTOMERS (العملاء) CRUD
# ═══════════════════════════════════════════════

@app.route('/customers')
@login_required
def customers_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()
    query = Customer.query.filter_by(user_id=current_user.id)
    if search:
        query = query.filter(
            or_(Customer.name.ilike(f'%{search}%'), Customer.phone.ilike(f'%{search}%'), Customer.email.ilike(f'%{search}%'))
        )
    pagination = query.order_by(Customer.name).paginate(page=page, per_page=PER_PAGE, error_out=False)
    return render_template('customers.html', user=current_user, customers=pagination.items, pagination=pagination)


@app.route('/customers/add', methods=['POST'])
@login_required
def customer_add():
    if not check_plan_limit('customers'):
        flash('لقد تجاوزت الحد المسموح به من العملاء في باقتك', 'danger')
        return redirect(url_for('customers_list'))
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم العميل مطلوب', 'danger')
        return redirect(url_for('customers_list'))
    customer = Customer(
        name=name,
        phone=request.form.get('phone', '').strip(),
        email=request.form.get('email', '').strip(),
        address=request.form.get('address', '').strip(),
        note=request.form.get('note', '').strip(),
        user_id=current_user.id,
    )
    db.session.add(customer)
    log_activity('create', 'عميل', customer.id, name)
    safe_commit()
    flash(f'تم إضافة العميل {name} بنجاح', 'success')
    return redirect(url_for('customers_list'))


@app.route('/customers/edit/<int:id>', methods=['POST'])
@login_required
def customer_edit(id):
    customer = Customer.query.get_or_404(id)
    if customer.user_id and customer.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم العميل مطلوب', 'danger')
        return redirect(url_for('customers_list'))
    customer.name = name
    customer.phone = request.form.get('phone', '').strip()
    customer.email = request.form.get('email', '').strip()
    customer.address = request.form.get('address', '').strip()
    customer.note = request.form.get('note', '').strip()
    log_activity('update', 'عميل', customer.id, name)
    safe_commit()
    flash(f'تم تحديث العميل {name} بنجاح', 'success')
    return redirect(url_for('customers_list'))


@app.route('/customers/delete/<int:id>', methods=['POST'])
@login_required
def customer_delete(id):
    customer = Customer.query.get_or_404(id)
    if customer.user_id and customer.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    name = customer.name
    db.session.delete(customer)
    log_activity('delete', 'عميل', id, name)
    safe_commit()
    flash(f'تم حذف العميل {name} بنجاح', 'success')
    return redirect(url_for('customers_list'))


# ═══════════════════════════════════════════════
# SUPPLIERS (الموردين) CRUD
# ═══════════════════════════════════════════════

@app.route('/suppliers')
@login_required
def suppliers_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()
    query = Supplier.query.filter_by(user_id=current_user.id)
    if search:
        query = query.filter(
            or_(Supplier.name.ilike(f'%{search}%'), Supplier.phone.ilike(f'%{search}%'))
        )
    pagination = query.order_by(Supplier.name).paginate(page=page, per_page=PER_PAGE, error_out=False)
    return render_template('suppliers.html', user=current_user, suppliers=pagination.items, pagination=pagination)


@app.route('/suppliers/add', methods=['POST'])
@login_required
def supplier_add():
    if not check_plan_limit('suppliers'):
        flash('لقد تجاوزت الحد المسموح به من الموردين في باقتك', 'danger')
        return redirect(url_for('suppliers_list'))
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم المورد مطلوب', 'danger')
        return redirect(url_for('suppliers_list'))
    supplier = Supplier(
        name=name,
        phone=request.form.get('phone', '').strip(),
        email=request.form.get('email', '').strip(),
        address=request.form.get('address', '').strip(),
        note=request.form.get('note', '').strip(),
        user_id=current_user.id,
    )
    db.session.add(supplier)
    log_activity('create', 'مورد', supplier.id, name)
    safe_commit()
    flash(f'تم إضافة المورد {name} بنجاح', 'success')
    return redirect(url_for('suppliers_list'))


@app.route('/suppliers/edit/<int:id>', methods=['POST'])
@login_required
def supplier_edit(id):
    supplier = Supplier.query.get_or_404(id)
    if supplier.user_id and supplier.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم المورد مطلوب', 'danger')
        return redirect(url_for('suppliers_list'))
    supplier.name = name
    supplier.phone = request.form.get('phone', '').strip()
    supplier.email = request.form.get('email', '').strip()
    supplier.address = request.form.get('address', '').strip()
    supplier.note = request.form.get('note', '').strip()
    log_activity('update', 'مورد', supplier.id, name)
    safe_commit()
    flash(f'تم تحديث المورد {name} بنجاح', 'success')
    return redirect(url_for('suppliers_list'))


@app.route('/suppliers/delete/<int:id>', methods=['POST'])
@login_required
def supplier_delete(id):
    supplier = Supplier.query.get_or_404(id)
    if supplier.user_id and supplier.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    name = supplier.name
    db.session.delete(supplier)
    log_activity('delete', 'مورد', id, name)
    safe_commit()
    flash(f'تم حذف المورد {name} بنجاح', 'success')
    return redirect(url_for('suppliers_list'))


# ═══════════════════════════════════════════════
# PURCHASES (المشتريات) CRUD
# ═══════════════════════════════════════════════

@app.route('/purchases')
@login_required
def purchases_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()
    query = Purchase.query.filter_by(user_id=current_user.id)
    if search:
        query = query.join(Purchase.supplier, isouter=True).filter(
            or_(Purchase.invoice_number.ilike(f'%{search}%'), Supplier.name.ilike(f'%{search}%'), Purchase.barcode.ilike(f'%{search}%'))
        )
    pagination = query.order_by(Purchase.created_at.desc()).paginate(page=page, per_page=PER_PAGE, error_out=False)
    return render_template('purchases.html', user=current_user, purchases=pagination.items, pagination=pagination)


@app.route('/purchases/add', methods=['GET', 'POST'])
@login_required
def purchase_add():
    if request.method == 'POST':
        if not check_plan_limit('purchases'):
            flash('لقد تجاوزت الحد المسموح به من فواتير المشتريات الشهرية في باقتك', 'danger')
            return redirect(url_for('purchases_list'))
        supplier_id = request.form.get('supplier_id', type=int)
        store_id = request.form.get('store_id', type=int)
        if not supplier_id or not store_id:
            flash('يرجى اختيار المورد والمخزن', 'danger')
            return redirect(url_for('purchases_list'))
        purchase = Purchase(
            invoice_number=generate_invoice('PUR'),
            supplier_id=supplier_id,
            store_id=store_id,
            user_id=current_user.id,
            total=request.form.get('total', 0, type=float),
            paid=request.form.get('paid', 0, type=float),
            note=request.form.get('note', '').strip(),
        )
        db.session.add(purchase)
        db.session.flush()
        purchase.barcode = f'2{str(purchase.id).zfill(10)}'

        item_ids = request.form.getlist('item_id[]')
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('price[]')
        for i in range(len(item_ids)):
            if item_ids[i] and quantities[i]:
                qty = int(quantities[i])
                price = float(prices[i]) if prices[i] else 0
                total = qty * price
                pi = PurchaseItem(
                    purchase_id=purchase.id,
                    item_id=int(item_ids[i]),
                    quantity=qty,
                    price=price,
                    total=total,
                )
                db.session.add(pi)
                item = db.session.get(Item, int(item_ids[i]))
                if item:
                    item.quantity = (item.quantity or 0) + qty

        log_activity('create', 'مشتريات', purchase.id, purchase.invoice_number)
        safe_commit()
        flash(f'تم إضافة فاتورة المشتريات {purchase.invoice_number} بنجاح', 'success')
        return redirect(url_for('purchases_list'))

    suppliers = Supplier.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    stores = Store.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    items = Item.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    us = get_user_settings()
    return render_template('purchase_form.html', user=current_user,
                           suppliers=suppliers, stores=stores, items=items,
                           purchase=None, us=us)


@app.route('/purchases/view/<int:id>')
@login_required
def purchase_view(id):
    purchase = Purchase.query.get_or_404(id)
    if purchase.user_id and purchase.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    suppliers = Supplier.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    stores = Store.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    items = Item.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    return render_template('purchase_form.html', user=current_user,
                           purchase=purchase, suppliers=suppliers,
                           stores=stores, items=items)


@app.route('/purchases/delete/<int:id>', methods=['POST'])
@login_required
def purchase_delete(id):
    purchase = Purchase.query.get_or_404(id)
    if purchase.user_id and purchase.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    inv = purchase.invoice_number
    for pi in purchase.items:
        item = db.session.get(Item, pi.item_id)
        if item:
            item.quantity = max(0, (item.quantity or 0) - pi.quantity)
    db.session.delete(purchase)
    log_activity('delete', 'مشتريات', id, inv)
    safe_commit()
    flash(f'تم حذف فاتورة المشتريات {inv} بنجاح', 'success')
    return redirect(url_for('purchases_list'))


@app.route('/purchases/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def purchase_edit(id):
    purchase = Purchase.query.get_or_404(id)
    if purchase.user_id and purchase.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    if request.method == 'POST':
        for pi in purchase.items:
            item = db.session.get(Item, pi.item_id)
            if item:
                item.quantity = max(0, (item.quantity or 0) - pi.quantity)
        purchase.supplier_id = request.form.get('supplier_id', type=int)
        purchase.store_id = request.form.get('store_id', type=int)
        purchase.total = request.form.get('total', 0, type=float)
        purchase.paid = request.form.get('paid', 0, type=float)
        purchase.note = request.form.get('note', '').strip()
        for pi in list(purchase.items):
            db.session.delete(pi)
        db.session.flush()
        item_ids = request.form.getlist('item_id[]')
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('price[]')
        for i in range(len(item_ids)):
            if item_ids[i] and quantities[i]:
                qty = int(quantities[i])
                price = float(prices[i]) if prices[i] else 0
                total = qty * price
                pi = PurchaseItem(
                    purchase_id=purchase.id,
                    item_id=int(item_ids[i]),
                    quantity=qty,
                    price=price,
                    total=total,
                )
                db.session.add(pi)
                item = db.session.get(Item, int(item_ids[i]))
                if item:
                    item.quantity = (item.quantity or 0) + qty
        log_activity('update', 'مشتريات', purchase.id, purchase.invoice_number)
        safe_commit()
        flash(f'تم تعديل فاتورة المشتريات {purchase.invoice_number} بنجاح', 'success')
        return redirect(url_for('purchases_list'))
    suppliers = Supplier.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    stores = Store.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    items = Item.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    return render_template('purchase_form.html', user=current_user,
                           suppliers=suppliers, stores=stores, items=items,
                           purchase=purchase, edit_mode=True)


# ═══════════════════════════════════════════════
# SALES (المبيعات) CRUD
# ═══════════════════════════════════════════════

@app.route('/sales')
@login_required
def sales_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()
    query = Sale.query.filter_by(user_id=current_user.id)
    if search:
        query = query.join(Sale.customer, isouter=True).filter(
            or_(Sale.invoice_number.ilike(f'%{search}%'), Customer.name.ilike(f'%{search}%'), Sale.barcode.ilike(f'%{search}%'))
        )
    pagination = query.order_by(Sale.created_at.desc()).paginate(page=page, per_page=PER_PAGE, error_out=False)
    return render_template('sales.html', user=current_user, sales=pagination.items, pagination=pagination)


@app.route('/sales/add', methods=['GET', 'POST'])
@login_required
def sale_add():
    if request.method == 'POST':
        if not check_plan_limit('sales'):
            flash('لقد تجاوزت الحد المسموح به من فواتير المبيعات الشهرية في باقتك', 'danger')
            return redirect(url_for('sales_list'))
        customer_id = request.form.get('customer_id', type=int)
        store_id = request.form.get('store_id', type=int)
        if not customer_id or not store_id:
            flash('يرجى اختيار العميل والمخزن', 'danger')
            return redirect(url_for('sales_list'))
        sale = Sale(
            invoice_number=generate_invoice('SALE'),
            customer_id=customer_id,
            store_id=store_id,
            user_id=current_user.id,
            total=request.form.get('total', 0, type=float),
            paid=request.form.get('paid', 0, type=float),
            discount=request.form.get('discount', 0, type=float),
            note=request.form.get('note', '').strip(),
        )
        db.session.add(sale)
        db.session.flush()
        sale.barcode = f'1{str(sale.id).zfill(10)}'

        item_ids = request.form.getlist('item_id[]')
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('price[]')
        for i in range(len(item_ids)):
            if item_ids[i] and quantities[i]:
                qty = int(quantities[i])
                price = float(prices[i]) if prices[i] else 0
                total = qty * price
                si = SaleItem(
                    sale_id=sale.id,
                    item_id=int(item_ids[i]),
                    quantity=qty,
                    price=price,
                    total=total,
                )
                db.session.add(si)
                item = db.session.get(Item, int(item_ids[i]))
                if item:
                    item.quantity = max(0, (item.quantity or 0) - qty)

        log_activity('create', 'مبيعات', sale.id, sale.invoice_number)
        safe_commit()
        flash(f'تم إضافة فاتورة المبيعات {sale.invoice_number} بنجاح', 'success')
        return redirect(url_for('sales_list'))

    customers = Customer.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    stores = Store.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    items = Item.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    us = get_user_settings()
    return render_template('sale_form.html', user=current_user,
                           customers=customers, stores=stores, items=items,
                           sale=None, us=us)


@app.route('/sales/view/<int:id>')
@login_required
def sale_view(id):
    sale = Sale.query.get_or_404(id)
    if sale.user_id and sale.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    customers = Customer.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    stores = Store.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    items = Item.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    return render_template('sale_form.html', user=current_user,
                           sale=sale, customers=customers,
                           stores=stores, items=items)


@app.route('/sales/delete/<int:id>', methods=['POST'])
@login_required
def sale_delete(id):
    sale = Sale.query.get_or_404(id)
    if sale.user_id and sale.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    inv = sale.invoice_number
    for si in sale.items:
        item = db.session.get(Item, si.item_id)
        if item:
            item.quantity = (item.quantity or 0) + si.quantity
    db.session.delete(sale)
    log_activity('delete', 'مبيعات', id, inv)
    safe_commit()
    flash(f'تم حذف فاتورة المبيعات {inv} بنجاح', 'success')
    return redirect(url_for('sales_list'))


@app.route('/sales/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def sale_edit(id):
    sale = Sale.query.get_or_404(id)
    if sale.user_id and sale.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    if request.method == 'POST':
        for si in sale.items:
            item = db.session.get(Item, si.item_id)
            if item:
                item.quantity = (item.quantity or 0) + si.quantity
        sale.customer_id = request.form.get('customer_id', type=int)
        sale.store_id = request.form.get('store_id', type=int)
        sale.total = request.form.get('total', 0, type=float)
        sale.paid = request.form.get('paid', 0, type=float)
        sale.discount = request.form.get('discount', 0, type=float)
        sale.note = request.form.get('note', '').strip()
        for si in list(sale.items):
            db.session.delete(si)
        db.session.flush()
        item_ids = request.form.getlist('item_id[]')
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('price[]')
        for i in range(len(item_ids)):
            if item_ids[i] and quantities[i]:
                qty = int(quantities[i])
                price = float(prices[i]) if prices[i] else 0
                total = qty * price
                si = SaleItem(
                    sale_id=sale.id,
                    item_id=int(item_ids[i]),
                    quantity=qty,
                    price=price,
                    total=total,
                )
                db.session.add(si)
                item = db.session.get(Item, int(item_ids[i]))
                if item:
                    item.quantity = max(0, (item.quantity or 0) - qty)
        log_activity('update', 'مبيعات', sale.id, sale.invoice_number)
        safe_commit()
        flash(f'تم تعديل فاتورة المبيعات {sale.invoice_number} بنجاح', 'success')
        return redirect(url_for('sales_list'))
    customers = Customer.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    stores = Store.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    items = Item.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    return render_template('sale_form.html', user=current_user,
                           customers=customers, stores=stores, items=items,
                           sale=sale, edit_mode=True)


# ═══════════════════════════════════════════════
# RETURNS (المرتجعات) CRUD
# ═══════════════════════════════════════════════

@app.route('/returns')
@login_required
def returns_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()
    query = Return.query.filter_by(user_id=current_user.id)
    if search:
        query = query.join(Return.sale, isouter=True).filter(
            or_(Return.return_number.ilike(f'%{search}%'), Sale.invoice_number.ilike(f'%{search}%'))
        )
    pagination = query.order_by(Return.created_at.desc()).paginate(page=page, per_page=PER_PAGE, error_out=False)
    return render_template('returns.html', user=current_user, returns=pagination.items, pagination=pagination)


@app.route('/returns/add', methods=['GET', 'POST'])
@login_required
def return_add():
    if request.method == 'POST':
        sale_id = request.form.get('sale_id', type=int)
        customer_id = request.form.get('customer_id', type=int)
        if not sale_id:
            flash('يرجى اختيار الفاتورة', 'danger')
            return redirect(url_for('returns_list'))
        ret = Return(
            return_number=generate_invoice('RET'),
            sale_id=sale_id,
            customer_id=customer_id,
            user_id=current_user.id,
            total=request.form.get('total', 0, type=float),
            reason=request.form.get('reason', '').strip(),
        )
        db.session.add(ret)
        db.session.flush()

        item_ids = request.form.getlist('item_id[]')
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('price[]')
        for i in range(len(item_ids)):
            if item_ids[i] and quantities[i]:
                qty = int(quantities[i])
                price = float(prices[i]) if prices[i] else 0
                total = qty * price
                ri = ReturnItem(
                    return_id=ret.id,
                    item_id=int(item_ids[i]),
                    quantity=qty,
                    price=price,
                    total=total,
                )
                db.session.add(ri)
                item = db.session.get(Item, int(item_ids[i]))
                if item:
                    item.quantity = (item.quantity or 0) + qty

        log_activity('create', 'مرتجع', ret.id, ret.return_number)
        safe_commit()
        flash(f'تم إضافة مرتجع {ret.return_number} بنجاح', 'success')
        return redirect(url_for('returns_list'))

    sales = Sale.query.filter_by(user_id=current_user.id).order_by(Sale.created_at.desc()).limit(DROPDOWN_LIMIT).all()
    customers = Customer.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    items = Item.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    return render_template('return_form.html', user=current_user,
                           sales=sales, customers=customers, items=items,
                           return_obj=None)


@app.route('/returns/view/<int:id>')
@login_required
def return_view(id):
    ret = Return.query.get_or_404(id)
    if ret.user_id and ret.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    sales = Sale.query.filter_by(user_id=current_user.id).order_by(Sale.created_at.desc()).limit(DROPDOWN_LIMIT).all()
    customers = Customer.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    items = Item.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    return render_template('return_form.html', user=current_user,
                           return_obj=ret, sales=sales,
                           customers=customers, items=items)


@app.route('/returns/delete/<int:id>', methods=['POST'])
@login_required
def return_delete(id):
    ret = Return.query.get_or_404(id)
    if ret.user_id and ret.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    rn = ret.return_number
    for ri in ret.items:
        item = db.session.get(Item, ri.item_id)
        if item:
            item.quantity = max(0, (item.quantity or 0) - ri.quantity)
    db.session.delete(ret)
    log_activity('delete', 'مرتجع', id, rn)
    safe_commit()
    flash(f'تم حذف المرتجع {rn} بنجاح', 'success')
    return redirect(url_for('returns_list'))


@app.route('/returns/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def return_edit(id):
    ret = Return.query.get_or_404(id)
    if ret.user_id and ret.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    if request.method == 'POST':
        for ri in ret.items:
            item = db.session.get(Item, ri.item_id)
            if item:
                item.quantity = max(0, (item.quantity or 0) - ri.quantity)
        ret.sale_id = request.form.get('sale_id', type=int)
        ret.customer_id = request.form.get('customer_id', type=int)
        ret.total = request.form.get('total', 0, type=float)
        ret.reason = request.form.get('reason', '').strip()
        for ri in list(ret.items):
            db.session.delete(ri)
        db.session.flush()
        item_ids = request.form.getlist('item_id[]')
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('price[]')
        for i in range(len(item_ids)):
            if item_ids[i] and quantities[i]:
                qty = int(quantities[i])
                price = float(prices[i]) if prices[i] else 0
                total = qty * price
                ri = ReturnItem(
                    return_id=ret.id,
                    item_id=int(item_ids[i]),
                    quantity=qty,
                    price=price,
                    total=total,
                )
                db.session.add(ri)
                item = db.session.get(Item, int(item_ids[i]))
                if item:
                    item.quantity = (item.quantity or 0) + qty
        log_activity('update', 'مرتجع', ret.id, ret.return_number)
        safe_commit()
        flash(f'تم تعديل المرتجع {ret.return_number} بنجاح', 'success')
        return redirect(url_for('returns_list'))
    sales = Sale.query.filter_by(user_id=current_user.id).order_by(Sale.created_at.desc()).limit(DROPDOWN_LIMIT).all()
    customers = Customer.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    items = Item.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    return render_template('return_form.html', user=current_user,
                           sales=sales, customers=customers, items=items,
                           return_obj=ret, edit_mode=True)


# ═══════════════════════════════════════════════
# EXPENSES (المصروفات) CRUD
# ═══════════════════════════════════════════════

@app.route('/expenses')
@login_required
def expenses_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()
    query = Expense.query.filter_by(user_id=current_user.id)
    if search:
        query = query.filter(Expense.description.ilike(f'%{search}%'))
    pagination = query.order_by(Expense.created_at.desc()).paginate(page=page, per_page=PER_PAGE, error_out=False)
    branches = Branch.query.filter_by(is_active=True, user_id=current_user.id).all()
    return render_template('expenses.html', user=current_user,
                           expenses=pagination.items, branches=branches, pagination=pagination)


@app.route('/expenses/add', methods=['POST'])
@login_required
def expense_add():
    description = request.form.get('description', '').strip()
    amount = request.form.get('amount', 0, type=float)
    if not description or amount <= 0:
        flash('التوصيف والمبلغ مطلوبان', 'danger')
        return redirect(url_for('expenses_list'))
    expense = Expense(
        description=description,
        amount=amount,
        category=request.form.get('category', '').strip(),
        branch_id=request.form.get('branch_id', type=int),
        user_id=current_user.id,
        note=request.form.get('note', '').strip(),
    )
    db.session.add(expense)
    log_activity('create', 'مصروف', expense.id, description)
    safe_commit()
    flash('تم إضافة المصروف بنجاح', 'success')
    return redirect(url_for('expenses_list'))


@app.route('/expenses/edit/<int:id>', methods=['POST'])
@login_required
def expense_edit(id):
    expense = Expense.query.get_or_404(id)
    if expense.user_id and expense.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    description = request.form.get('description', '').strip()
    amount = request.form.get('amount', 0, type=float)
    if not description or amount <= 0:
        flash('التوصيف والمبلغ مطلوبان', 'danger')
        return redirect(url_for('expenses_list'))
    expense.description = description
    expense.amount = amount
    expense.category = request.form.get('category', '').strip()
    expense.branch_id = request.form.get('branch_id', type=int)
    expense.note = request.form.get('note', '').strip()
    log_activity('update', 'مصروف', expense.id, description)
    safe_commit()
    flash('تم تحديث المصروف بنجاح', 'success')
    return redirect(url_for('expenses_list'))


@app.route('/expenses/delete/<int:id>', methods=['POST'])
@login_required
def expense_delete(id):
    expense = Expense.query.get_or_404(id)
    if expense.user_id and expense.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    db.session.delete(expense)
    log_activity('delete', 'مصروف', expense.id, expense.description)
    safe_commit()
    flash('تم حذف المصروف بنجاح', 'success')
    return redirect(url_for('expenses_list'))


# ═══════════════════════════════════════════════
# BONDS (السندات) CRUD
# ═══════════════════════════════════════════════

@app.route('/bonds')
@login_required
def bonds_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()
    query = Bond.query.filter_by(user_id=current_user.id)
    if search:
        query = query.filter(Bond.bond_number.ilike(f'%{search}%'))
    pagination = query.order_by(Bond.created_at.desc()).paginate(page=page, per_page=PER_PAGE, error_out=False)
    customers = Customer.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    suppliers = Supplier.query.filter_by(is_active=True, user_id=current_user.id).limit(DROPDOWN_LIMIT).all()
    return render_template('bonds.html', user=current_user, bonds=pagination.items,
                           customers=customers, suppliers=suppliers, pagination=pagination)


@app.route('/bonds/add', methods=['POST'])
@login_required
def bond_add():
    bond_type = request.form.get('bond_type', '')
    amount = request.form.get('amount', 0, type=float)
    if not bond_type or amount <= 0:
        flash('نوع السند والمبلغ مطلوبان', 'danger')
        return redirect(url_for('bonds_list'))
    bond = Bond(
        bond_number=generate_invoice('BND'),
        bond_type=bond_type,
        amount=amount,
        customer_id=request.form.get('customer_id', type=int),
        supplier_id=request.form.get('supplier_id', type=int),
        user_id=current_user.id,
        note=request.form.get('note', '').strip(),
    )
    db.session.add(bond)
    log_activity('create', 'سند', bond.id, bond.bond_number)
    safe_commit()
    flash(f'تم إضافة السند {bond.bond_number} بنجاح', 'success')
    return redirect(url_for('bonds_list'))


@app.route('/bonds/delete/<int:id>', methods=['POST'])
@login_required
def bond_delete(id):
    bond = Bond.query.get_or_404(id)
    if bond.user_id and bond.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    bn = bond.bond_number
    db.session.delete(bond)
    log_activity('delete', 'سند', id, bn)
    safe_commit()
    flash(f'تم حذف السند {bn} بنجاح', 'success')
    return redirect(url_for('bonds_list'))


@app.route('/bonds/edit/<int:id>', methods=['POST'])
@login_required
def bond_edit(id):
    bond = Bond.query.get_or_404(id)
    if bond.user_id and bond.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    bond_type = request.form.get('bond_type', '')
    amount = request.form.get('amount', 0, type=float)
    if not bond_type or amount <= 0:
        flash('نوع السند والمبلغ مطلوبان', 'danger')
        return redirect(url_for('bonds_list'))
    bond.bond_type = bond_type
    bond.amount = amount
    bond.customer_id = request.form.get('customer_id', type=int)
    bond.supplier_id = request.form.get('supplier_id', type=int)
    bond.note = request.form.get('note', '').strip()
    log_activity('update', 'سند', bond.id, bond.bond_number)
    safe_commit()
    flash(f'تم تعديل السند {bond.bond_number} بنجاح', 'success')
    return redirect(url_for('bonds_list'))


# ═══════════════════════════════════════════════
# INVOICE SETTINGS & PRINT
# ═══════════════════════════════════════════════

def get_store_settings():
    settings = StoreSetting.query.filter_by(user_id=current_user.id).first()
    if not settings:
        settings = StoreSetting(user_id=current_user.id)
        db.session.add(settings)
        safe_commit()
    return settings


# Make datetime available in all templates
@app.context_processor
def inject_globals():
    extra = {'datetime': datetime, 'now': datetime.utcnow(), 'Employee': None, 'current_employee': None, 'employee_permissions': {}}
    if current_user.is_authenticated and current_user.role == 'employee':
        emp = Employee.query.filter_by(user_id=current_user.id, is_active=True).first()
        if emp:
            extra['current_employee'] = emp
            extra['employee_permissions'] = emp.get_permissions()
    # Inject Employee model for sidebar queries
    from models import Employee as EmpModel
    extra['Employee'] = EmpModel
    return extra


_orig_commit = db.session.commit


def log_activity(action, entity_type, entity_id=None, details=None):
    if not current_user.is_authenticated or current_user.is_super_admin:
        return
    log = ActivityLog(
        user_id=current_user.id,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details,
    )
    db.session.add(log)


def safe_commit():
    """Replace safe_commit() with safe_commit() for automatic error handling + rollback."""
    try:
        _orig_commit()
    except Exception:
        db.session.rollback()


@app.errorhandler(500)
def internal_error(e):
    db.session.rollback()
    if current_user.is_authenticated:
        flash('عذراً، حدث خطأ غير متوقع. تم إلغاء العملية.', 'danger')
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.before_request
def block_unauthenticated_and_employees():
    if current_user.is_authenticated:
        ep = request.endpoint or ''
        is_api = ep.startswith('api.')

        # Super admin: only access admin pages + API
        if current_user.is_super_admin:
            allowed = ['admin_dashboard', 'admin_users', 'admin_plans', 'admin_subscriptions',
                       'admin_notifications', 'admin_send_notification',
                       'admin_toggle_user', 'admin_delete_user', 'admin_subscribe_user',
                       'admin_add_plan', 'admin_edit_plan', 'admin_toggle_plan',
                       'admin_toggle_subscription', 'user_guide', 'support', 'landing', 'login',
                       'logout', 'static', 'check_auth', 'notifications_list', 'notifications_count',
                       'notification_mark_read', 'notification_mark_all_read']
            if not is_api and ep and ep not in allowed and not ep.startswith('admin_'):
                flash('صفحات النظام متاحة فقط للمستخدمين العاديين', 'warning')
                return redirect(url_for('admin_dashboard'))

        # Subscription check: allow API, skip for super admin
        if not current_user.is_super_admin and not current_user.is_subscription_active:
            if not is_api and ep and ep not in ('login', 'logout', 'static', 'landing', 'check_auth', 'subscription_expired'):
                return redirect(url_for('subscription_expired'))

        # Employee: check page permissions (skip API)
        if current_user.role == 'employee':
            if is_api:
                pass  # API routes use @api_login_required decorator instead
            else:
                employee = Employee.query.filter_by(user_id=current_user.id, is_active=True).first()
                if not employee:
                    flash('ليس لديك صلاحية للوصول إلى النظام', 'danger')
                    logout_user()
                    return redirect(url_for('login'))

                allowed_endpoints = [
                    'login', 'logout', 'static', 'landing', 'check_auth', 'user_guide',
                    'dashboard', 'subscription_expired', 'notifications_list',
                    'notifications_count', 'notification_mark_read', 'notification_mark_all_read'
                ]
                if ep and ep not in allowed_endpoints:
                    if not employee.has_perm(ep):
                        redirect_target = find_first_available_page(employee)
                        if redirect_target:
                            flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'warning')
                            return redirect(url_for(redirect_target))
                        logout_user()
                        flash('ليس لديك أي صلاحيات في النظام', 'danger')
                        return redirect(url_for('login'))


def find_first_available_page(employee):
    endpoints = [
        'branches_list', 'stores_list', 'categories_list', 'items_list',
        'customers_list', 'suppliers_list', 'purchases_list', 'sales_list',
        'returns_list', 'expenses_list', 'bonds_list',
        'user_guide', 'support',
    ]
    for ep in endpoints:
        if employee.has_perm(ep):
            return ep
    if employee.has_perm('dashboard'):
        return 'dashboard'
    return None


def check_plan_limit(resource_type):
    """Check if current user's plan allows adding more of a resource."""
    if current_user.is_super_admin:
        return True
    sub = current_user.subscription
    if not sub or not sub.is_active:
        return False
    plan = sub.plan
    if not plan:
        return False
    now = datetime.utcnow()
    month_start = datetime(now.year, now.month, 1)
    if resource_type == 'items':
        count = Item.query.filter_by(is_active=True, user_id=current_user.id).count()
        return count < plan.max_items
    elif resource_type == 'customers':
        count = Customer.query.filter_by(is_active=True, user_id=current_user.id).count()
        return count < plan.max_customers
    elif resource_type == 'suppliers':
        count = Supplier.query.filter_by(is_active=True, user_id=current_user.id).count()
        return count < plan.max_suppliers
    elif resource_type == 'sales':
        count = Sale.query.filter(Sale.created_at >= month_start, Sale.user_id == current_user.id).count()
        return count < plan.max_invoices_monthly
    elif resource_type == 'purchases':
        count = Purchase.query.filter(Purchase.created_at >= month_start, Purchase.user_id == current_user.id).count()
        return count < plan.max_invoices_monthly
    elif resource_type == 'branches':
        count = Branch.query.filter_by(is_active=True, user_id=current_user.id).count()
        return count < plan.max_branches
    elif resource_type == 'stores':
        count = Store.query.filter_by(is_active=True, user_id=current_user.id).count()
        return count < plan.max_stores
    return True


def regular_user_required(f):
    """Redirect super admins to admin dashboard if they try to access store pages."""
    from functools import wraps
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.is_super_admin:
            flash('صفحات النظام متاحة فقط للمستخدمين العاديين', 'warning')
            return redirect(url_for('admin_dashboard'))
        return f(*args, **kwargs)
    return decorated


@app.route('/invoice-settings', methods=['GET', 'POST'])
@login_required
def invoice_settings():
    settings = get_store_settings()
    if request.method == 'POST':
        settings.store_name = request.form.get('store_name', 'Mazad Plus').strip()
        settings.store_phone = request.form.get('store_phone', '').strip()
        settings.store_address = request.form.get('store_address', '').strip()
        settings.store_email = request.form.get('store_email', '').strip()
        settings.tax_number = request.form.get('tax_number', '').strip()
        settings.invoice_template = request.form.get('invoice_template', 1, type=int)
        settings.invoice_footer = request.form.get('invoice_footer', '').strip()
        settings.currency = request.form.get('currency', 'ج.م').strip()
        log_activity('update', 'إعدادات الفاتورة', settings.id, settings.store_name)
        safe_commit()
        flash('تم حفظ إعدادات الفواتير بنجاح', 'success')
        return redirect(url_for('invoice_settings'))
    return render_template('invoice_settings.html', user=current_user, settings=settings)


def get_user_settings():
    us = UserSetting.query.filter_by(user_id=current_user.id).first()
    if not us:
        us = UserSetting(user_id=current_user.id)
        db.session.add(us)
        safe_commit()
    return us


@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    us = get_user_settings()
    if request.method == 'POST':
        us.language = request.form.get('language', 'ar')
        us.currency = request.form.get('currency', 'ج.م').strip()
        us.print_copies = request.form.get('print_copies', 1, type=int)
        us.page_size = request.form.get('page_size', 20, type=int)
        us.default_payment_status = request.form.get('default_payment_status', 'unpaid')
        us.notifications_enabled = request.form.get('notifications_enabled') == 'on'
        us.dark_mode = request.form.get('dark_mode') == 'on'
        safe_commit()
        flash('تم حفظ الإعدادات بنجاح', 'success')
        return redirect(url_for('settings'))
    return render_template('settings.html', user=current_user, us=us)


UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/upload-logo', methods=['POST'])
@login_required
def upload_logo():
    if 'logo' not in request.files:
        flash('لم يتم تحديد ملف', 'danger')
        return redirect(url_for('invoice_settings'))
    file = request.files['logo']
    if file.filename == '':
        flash('لم يتم تحديد ملف', 'danger')
        return redirect(url_for('invoice_settings'))
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'png'
    filename = f'logo_{current_user.id}.{ext}'
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)
    settings = get_store_settings()
    settings.store_logo = f'/static/uploads/{filename}'
    safe_commit()
    flash('تم رفع الشعار بنجاح', 'success')
    return redirect(url_for('invoice_settings'))


@app.route('/remove-logo', methods=['POST'])
@login_required
def remove_logo():
    settings = get_store_settings()
    settings.store_logo = None
    safe_commit()
    flash('تم إزالة الشعار', 'success')
    return redirect(url_for('invoice_settings'))


@app.route('/sale/print/<int:id>/<int:template_id>')
@login_required
def sale_print(id, template_id):
    sale = Sale.query.get_or_404(id)
    if sale.user_id and sale.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    settings = get_store_settings()
    if template_id < 1 or template_id > 3:
        template_id = settings.invoice_template
    template_name = f'invoice_templates/template{template_id}.html'
    return render_template(template_name, sale=sale, settings=settings, user=current_user)


@app.route('/purchase/print/<int:id>/<int:template_id>')
@login_required
def purchase_print(id, template_id):
    purchase = Purchase.query.get_or_404(id)
    if purchase.user_id and purchase.user_id != current_user.id and not current_user.is_super_admin:
        abort(403)
    settings = get_store_settings()
    if template_id < 1 or template_id > 3:
        template_id = settings.invoice_template
    return render_template(f'invoice_templates/purchase_template{template_id}.html',
                           purchase=purchase, settings=settings, user=current_user)


# ═══════════════════════════════════════════════
# USER GUIDE
# ═══════════════════════════════════════════════

@app.route('/user-guide')
@login_required
def user_guide():
    return render_template('user_guide.html', user=current_user)


@app.route('/subscription-expired')
@login_required
def subscription_expired():
    sub = current_user.subscription
    return render_template('subscription_expired.html', user=current_user, sub=sub)


@app.route('/support')
def support():
    return render_template('support.html')


@app.route('/about')
def about():
    return render_template('about.html')


@app.route('/features')
def features():
    return render_template('features.html')


@app.route('/privacy')
def privacy():
    return render_template('privacy.html')


@app.route('/terms')
def terms():
    return render_template('terms.html')


@app.route('/download')
def download():
    return render_template('download.html')


# ═══════════════════════════════════════════════
# SUPER ADMIN ROUTES
# ═══════════════════════════════════════════════

def admin_required(f):
    from functools import wraps
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_super_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated


@app.route('/admin')
@admin_required
def admin_dashboard():
    page = request.args.get('page', 1, type=int)
    users_pagination = User.query.order_by(User.created_at.desc()).paginate(page=page, per_page=PER_PAGE, error_out=False)
    users = users_pagination.items
    plans = Plan.query.all()
    active_subs = UserSubscription.query.filter_by(is_active=True).count()
    total_revenue = db.session.query(db.func.sum(Plan.price)).filter(
        Plan.id.in_([s.plan_id for s in UserSubscription.query.filter_by(is_active=True).all()])
    ).scalar() or 0
    return render_template('admin_dashboard.html', user=current_user,
                           users=users, plans=plans,
                           active_subs=active_subs,
                           total_revenue=total_revenue,
                           stats={
                               'total_users': User.query.count(),
                               'total_plans': Plan.query.count(),
                               'total_subs': UserSubscription.query.count(),
                               'active_subs': active_subs,
                           })


@app.route('/admin/users')
@admin_required
def admin_users():
    page = request.args.get('page', 1, type=int)
    pagination = User.query.order_by(User.created_at.desc()).paginate(page=page, per_page=PER_PAGE, error_out=False)
    plans = Plan.query.filter_by(is_active=True).all()
    return render_template('admin_users.html', user=current_user,
                           users=pagination.items, plans=plans, pagination=pagination)


@app.route('/admin/users/toggle/<int:id>', methods=['POST'])
@admin_required
def admin_toggle_user(id):
    target = User.query.get_or_404(id)
    if target.is_super_admin:
        flash('لا يمكن تعطيل حساب السوبر أدمن', 'danger')
        return redirect(url_for('admin_users'))
    target.is_active = not target.is_active
    log_activity('update', 'مستخدم', target.id, target.full_name)
    safe_commit()
    flash(f'تم {"تفعيل" if target.is_active else "تعطيل"} حساب {target.full_name}', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/delete/<int:id>', methods=['POST'])
@admin_required
def admin_delete_user(id):
    target = User.query.get_or_404(id)
    if target.is_super_admin:
        flash('لا يمكن حذف حساب السوبر أدمن', 'danger')
        return redirect(url_for('admin_users'))
    name = target.full_name
    # Clean up related data
    StoreSetting.query.filter_by(user_id=id).delete()
    UserSubscription.query.filter_by(user_id=id).delete()
    db.session.delete(target)
    log_activity('delete', 'مستخدم', id, name)
    safe_commit()
    flash(f'تم حذف المستخدم {name}', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/subscribe/<int:id>', methods=['POST'])
@admin_required
def admin_subscribe_user(id):
    target = User.query.get_or_404(id)
    plan_id = request.form.get('plan_id', type=int)
    duration = request.form.get('duration', 30, type=int)
    plan = Plan.query.get_or_404(plan_id)
    sub = UserSubscription(
        user_id=id,
        plan_id=plan_id,
        start_date=datetime.utcnow(),
        end_date=datetime.utcnow() + timedelta(days=duration),
        is_active=True,
        payment_status='paid',
        payment_method='admin',
        notes=f'اشتراك يدوي بواسطة {current_user.full_name}',
    )
    db.session.add(sub)
    log_activity('create', 'اشتراك', sub.id, f'{target.full_name} -> {plan.name}')
    safe_commit()
    flash(f'تم تفعيل اشتراك {target.full_name} في خطة {plan.name}', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/plans')
@admin_required
def admin_plans():
    all_plans = Plan.query.order_by(Plan.price).all()
    return render_template('admin_plans.html', user=current_user, plans=all_plans)


@app.route('/admin/plans/add', methods=['POST'])
@admin_required
def admin_add_plan():
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم الخطة مطلوب', 'danger')
        return redirect(url_for('admin_plans'))
    plan = Plan(
        name=name,
        description=request.form.get('description', '').strip(),
        price=request.form.get('price', 0, type=float),
        duration_days=request.form.get('duration_days', 30, type=int),
        trial_days=request.form.get('trial_days', 7, type=int),
        max_branches=request.form.get('max_branches', 1, type=int),
        max_stores=request.form.get('max_stores', 1, type=int),
        max_items=request.form.get('max_items', 50, type=int),
        max_customers=request.form.get('max_customers', 50, type=int),
        max_suppliers=request.form.get('max_suppliers', 20, type=int),
        max_invoices_monthly=request.form.get('max_invoices_monthly', 100, type=int),
        max_users=request.form.get('max_users', 1, type=int),
        features=request.form.get('features', '').strip(),
    )
    db.session.add(plan)
    log_activity('create', 'خطة', plan.id, name)
    safe_commit()
    flash(f'تم إضافة الخطة {name} بنجاح', 'success')
    return redirect(url_for('admin_plans'))


@app.route('/admin/plans/edit/<int:id>', methods=['POST'])
@admin_required
def admin_edit_plan(id):
    plan = Plan.query.get_or_404(id)
    name = request.form.get('name', '').strip()
    if not name:
        flash('اسم الخطة مطلوب', 'danger')
        return redirect(url_for('admin_plans'))
    plan.name = name
    plan.description = request.form.get('description', '').strip()
    plan.price = request.form.get('price', 0, type=float)
    plan.duration_days = request.form.get('duration_days', 30, type=int)
    plan.trial_days = request.form.get('trial_days', 7, type=int)
    plan.max_branches = request.form.get('max_branches', 1, type=int)
    plan.max_stores = request.form.get('max_stores', 1, type=int)
    plan.max_items = request.form.get('max_items', 50, type=int)
    plan.max_customers = request.form.get('max_customers', 50, type=int)
    plan.max_suppliers = request.form.get('max_suppliers', 20, type=int)
    plan.max_invoices_monthly = request.form.get('max_invoices_monthly', 100, type=int)
    plan.max_users = request.form.get('max_users', 1, type=int)
    plan.features = request.form.get('features', '').strip()
    log_activity('update', 'خطة', plan.id, name)
    safe_commit()
    flash(f'تم تحديث الخطة {name} بنجاح', 'success')
    return redirect(url_for('admin_plans'))


@app.route('/admin/plans/toggle/<int:id>', methods=['POST'])
@admin_required
def admin_toggle_plan(id):
    plan = Plan.query.get_or_404(id)
    plan.is_active = not plan.is_active
    log_activity('update', 'خطة', plan.id, plan.name)
    safe_commit()
    flash(f'تم {"تفعيل" if plan.is_active else "تعطيل"} الخطة {plan.name}', 'success')
    return redirect(url_for('admin_plans'))


@app.route('/admin/subscriptions')
@admin_required
def admin_subscriptions():
    subs = UserSubscription.query.order_by(UserSubscription.created_at.desc()).all()
    return render_template('admin_subscriptions.html', user=current_user, subs=subs)


@app.route('/admin/subscriptions/toggle/<int:id>', methods=['POST'])
@admin_required
def admin_toggle_subscription(id):
    sub = UserSubscription.query.get_or_404(id)
    sub.is_active = not sub.is_active
    log_activity('update', 'اشتراك', sub.id, f'اشتراك {sub.id}')
    safe_commit()
    flash(f'تم {"تفعيل" if sub.is_active else "إيقاف"} الاشتراك', 'success')
    return redirect(url_for('admin_subscriptions'))


# ─────────────────────────── Employee Management ───────────────────────────

def owner_required(f):
    from functools import wraps
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.is_super_admin:
            return redirect(url_for('admin_dashboard'))
        # Employees cannot manage other employees
        if current_user.role == 'employee':
            flash('ليس لديك صلاحية لإدارة الموظفين', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


ALL_PAGES = [
    ('dashboard', 'لوحة التحكم', 'fas fa-chart-pie'),
    ('branches_list', 'الفروع', 'fas fa-code-branch'),
    ('branch_add', 'إضافة فرع', 'fas fa-plus-circle'),
    ('stores_list', 'المخازن', 'fas fa-warehouse'),
    ('store_add', 'إضافة مخزن', 'fas fa-plus-circle'),
    ('categories_list', 'التصنيفات', 'fas fa-tags'),
    ('category_add', 'إضافة تصنيف', 'fas fa-plus-circle'),
    ('items_list', 'الأصناف', 'fas fa-boxes'),
    ('item_add', 'إضافة صنف', 'fas fa-plus-circle'),
    ('customers_list', 'العملاء', 'fas fa-users'),
    ('customer_add', 'إضافة عميل', 'fas fa-plus-circle'),
    ('suppliers_list', 'الموردين', 'fas fa-truck'),
    ('supplier_add', 'إضافة مورد', 'fas fa-plus-circle'),
    ('purchases_list', 'المشتريات', 'fas fa-shopping-cart'),
    ('purchase_add', 'إضافة فاتورة مشتريات', 'fas fa-plus-circle'),
    ('purchase_edit', 'تعديل فاتورة مشتريات', 'fas fa-edit'),
    ('sales_list', 'المبيعات', 'fas fa-cash-register'),
    ('sale_add', 'إضافة فاتورة مبيعات', 'fas fa-plus-circle'),
    ('sale_edit', 'تعديل فاتورة مبيعات', 'fas fa-edit'),
    ('returns_list', 'المرتجعات', 'fas fa-undo-alt'),
    ('return_add', 'إضافة مرتجع', 'fas fa-plus-circle'),
    ('return_edit', 'تعديل مرتجع', 'fas fa-edit'),
    ('expenses_list', 'المصروفات', 'fas fa-money-bill-wave'),
    ('expense_add', 'إضافة مصروف', 'fas fa-plus-circle'),
    ('expense_edit', 'تعديل مصروف', 'fas fa-edit'),
    ('bonds_list', 'السندات', 'fas fa-file-invoice'),
    ('bond_add', 'إضافة سند', 'fas fa-plus-circle'),
    ('bond_edit', 'تعديل سند', 'fas fa-edit'),
    ('invoice_settings', 'إعدادات الفواتير', 'fas fa-cog'),
    ('settings', 'الإعدادات العامة', 'fas fa-sliders-h'),
    ('user_guide', 'دليل الاستخدام', 'fas fa-book'),
    ('support', 'الدعم الفني', 'fas fa-headset'),
]


@app.route('/employees')
@owner_required
def employees():
    employees_list = Employee.query.filter_by(added_by=current_user.id).order_by(Employee.created_at.desc()).all()
    branches = Branch.query.filter_by(is_active=True, user_id=current_user.id).all()

    import json
    employees_json = []
    for e in employees_list:
        employees_json.append({
            'id': e.id,
            'user_id': e.user_id,
            'added_by': e.added_by,
            'branch_id': e.branch_id,
            'phone': e.phone,
            'permissions': e.permissions,
            'is_active': e.is_active,
            'created_at': e.created_at.isoformat() if e.created_at else None,
            'user': {
                'id': e.user.id if e.user else None,
                'full_name': e.user.full_name if e.user else 'موظف',
                'email': e.user.email if e.user else '',
                'phone': e.user.phone if e.user else '',
                'is_active': e.user.is_active if e.user else False,
            } if e.user else None,
            'branch': {
                'id': e.branch.id if e.branch else None,
                'name': e.branch.name if e.branch else 'جميع الفروع',
            } if e.branch else None,
        })

    branches_json = [{'id': b.id, 'name': b.name} for b in branches]

    return render_template('employees.html', user=current_user,
                         employees=employees_list, employees_json=json.dumps(employees_json, ensure_ascii=False),
                         branches=branches, branches_json=json.dumps(branches_json, ensure_ascii=False),
                         all_pages=ALL_PAGES)


@app.route('/employees/add', methods=['POST'])
@owner_required
def employee_add():
    full_name = request.form.get('full_name', '').strip()
    email = request.form.get('email', '').strip()
    phone = request.form.get('phone', '').strip()
    password = request.form.get('password', '').strip()
    branch_id = request.form.get('branch_id', type=int)

    if not full_name or not email or not password:
        flash('الاسم والبريد الإلكتروني وكلمة المرور مطلوبة', 'danger')
        return redirect(url_for('employees'))

    if User.query.filter_by(email=email).first():
        flash('هذا البريد الإلكتروني مستخدم بالفعل', 'danger')
        return redirect(url_for('employees'))

    # Create user account for employee
    emp_user = User(full_name=full_name, email=email, phone=phone, role='employee')
    emp_user.set_password(password)
    db.session.add(emp_user)
    db.session.flush()

    # Collect permissions
    permissions = {}
    for ep, _, _ in ALL_PAGES:
        permissions[ep] = request.form.get(f'perm_{ep}') == 'on'

    employee = Employee(
        user_id=emp_user.id,
        added_by=current_user.id,
        branch_id=branch_id,
        phone=phone,
        permissions=__import__('json').dumps(permissions, ensure_ascii=False),
    )
    db.session.add(employee)
    log_activity('create', 'موظف', employee.id, full_name)
    safe_commit()

    flash(f'تم إضافة الموظف {full_name} بنجاح', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/edit/<int:id>', methods=['POST'])
@owner_required
def employee_edit(id):
    employee = Employee.query.get_or_404(id)
    if employee.added_by != current_user.id:
        flash('لا يمكنك تعديل هذا الموظف', 'danger')
        return redirect(url_for('employees'))

    emp_user = employee.user
    full_name = request.form.get('full_name', '').strip()
    email = request.form.get('email', '').strip()
    phone = request.form.get('phone', '').strip()
    password = request.form.get('password', '').strip()
    branch_id = request.form.get('branch_id', type=int)

    if not full_name or not email:
        flash('الاسم والبريد الإلكتروني مطلوبان', 'danger')
        return redirect(url_for('employees'))

    # Check email uniqueness
    existing = User.query.filter_by(email=email).first()
    if existing and existing.id != emp_user.id:
        flash('هذا البريد الإلكتروني مستخدم بالفعل', 'danger')
        return redirect(url_for('employees'))

    emp_user.full_name = full_name
    emp_user.email = email
    emp_user.phone = phone
    if password:
        emp_user.set_password(password)

    employee.phone = phone
    employee.branch_id = branch_id

    permissions = {}
    for ep, _, _ in ALL_PAGES:
        permissions[ep] = request.form.get(f'perm_{ep}') == 'on'
    employee.permissions = __import__('json').dumps(permissions, ensure_ascii=False)

    log_activity('update', 'موظف', employee.id, full_name)
    safe_commit()
    flash(f'تم تعديل بيانات الموظف {full_name}', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/delete/<int:id>', methods=['POST'])
@owner_required
def employee_delete(id):
    employee = Employee.query.get_or_404(id)
    if employee.added_by != current_user.id:
        flash('لا يمكنك حذف هذا الموظف', 'danger')
        return redirect(url_for('employees'))

    emp_user = employee.user
    # Deactivate user instead of deleting
    if emp_user:
        emp_user.is_active = False
    employee.is_active = False
    log_activity('delete', 'موظف', employee.id, employee.user.full_name if employee.user else '')
    safe_commit()
    flash('تم تعطيل حساب الموظف', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/toggle/<int:id>', methods=['POST'])
@owner_required
def employee_toggle(id):
    employee = Employee.query.get_or_404(id)
    if employee.added_by != current_user.id:
        flash('لا يمكنك تعديل هذا الموظف', 'danger')
        return redirect(url_for('employees'))

    employee.is_active = not employee.is_active
    if employee.user:
        employee.user.is_active = employee.is_active
    log_activity('update', 'موظف', employee.id, employee.user.full_name if employee.user else '')
    safe_commit()
    status = 'تفعيل' if employee.is_active else 'إيقاف'
    flash(f'تم {status} حساب الموظف', 'success')
    return redirect(url_for('employees'))


@app.route('/api/sale-items/<int:sale_id>')
@csrf.exempt
@login_required
def api_sale_items(sale_id):
    sale = Sale.query.get_or_404(sale_id)
    if sale.user_id and sale.user_id != current_user.id and not current_user.is_super_admin:
        return jsonify({'error': 'غير مصرح'}), 403
    items = [{
        'item_name': si.item.name if si.item else 'منتج',
        'quantity': si.quantity,
        'price': si.price,
        'total': si.total,
    } for si in sale.items]
    return jsonify({
        'invoice_number': sale.invoice_number,
        'customer_name': sale.customer.name if sale.customer else 'عميل نقدي',
        'customer_phone': sale.customer.phone if sale.customer else '',
        'customer_address': sale.customer.address if sale.customer else '',
        'date': sale.created_at.strftime('%Y-%m-%d %H:%M'),
        'total': sale.total,
        'discount': sale.discount,
        'paid': sale.paid,
        'remaining': sale.remaining,
        'status': sale.status,
        'items': items,
    })


# ═══════════════════════════════════════════════
# DATA EXPORT (Excel & PDF)
# ═══════════════════════════════════════════════

EXPORT_CONFIG = {
    'branches': {
        'title': 'الفروع',
        'columns': ['#', 'الاسم', 'المدير', 'الهاتف', 'العنوان', 'الحالة'],
        'data_fn': lambda: Branch.query.order_by(Branch.name).all(),
        'row_fn': lambda r: [r.id, r.name, r.manager or '-', r.phone or '-', r.address or '-', 'نشط' if r.is_active else 'غير نشط'],
    },
    'stores': {
        'title': 'المخازن',
        'columns': ['#', 'الاسم', 'الفرع', 'المدير', 'العنوان', 'الحالة'],
        'data_fn': lambda: Store.query.order_by(Store.name).all(),
        'row_fn': lambda r: [r.id, r.name, r.branch.name if r.branch else '-', r.manager or '-', r.address or '-', 'نشط' if r.is_active else 'غير نشط'],
    },
    'categories': {
        'title': 'التصنيفات',
        'columns': ['#', 'الاسم', 'الوصف', 'عدد الأصناف', 'الحالة'],
        'data_fn': lambda: Category.query.order_by(Category.name).all(),
        'row_fn': lambda r: [r.id, r.name, r.description or '-', len(r.items) if r.items else 0, 'نشط' if r.is_active else 'غير نشط'],
    },
    'items': {
        'title': 'الأصناف',
        'columns': ['#', 'الاسم', 'الباركود', 'التصنيف', 'المخزن', 'سعر الشراء', 'سعر البيع', 'الكمية', 'الحالة'],
        'data_fn': lambda: Item.query.order_by(Item.name).all(),
        'row_fn': lambda r: [r.id, r.name, r.barcode or '-', r.category.name if r.category else '-', r.store.name if r.store else '-', r.buy_price, r.sell_price, r.quantity, 'منخفض' if r.stock_status == 'low' else 'نافد' if r.stock_status == 'out' else 'متوفر'],
    },
    'customers': {
        'title': 'العملاء',
        'columns': ['#', 'الاسم', 'الهاتف', 'البريد', 'العنوان', 'إجمالي المشتريات', 'المدفوع', 'المتبقي', 'الحالة'],
        'data_fn': lambda: Customer.query.order_by(Customer.name).all(),
        'row_fn': lambda r: [r.id, r.name, r.phone or '-', r.email or '-', r.address or '-', r.total_purchases, r.total_paid, r.balance, 'نشط' if r.is_active else 'غير نشط'],
    },
    'suppliers': {
        'title': 'الموردين',
        'columns': ['#', 'الاسم', 'الهاتف', 'البريد', 'العنوان', 'ملاحظات', 'الحالة'],
        'data_fn': lambda: Supplier.query.order_by(Supplier.name).all(),
        'row_fn': lambda r: [r.id, r.name, r.phone or '-', r.email or '-', r.address or '-', r.note or '-', 'نشط' if r.is_active else 'غير نشط'],
    },
    'purchases': {
        'title': 'المشتريات',
        'columns': ['#', 'رقم الفاتورة', 'المورد', 'المخزن', 'الإجمالي', 'المدفوع', 'المتبقي', 'الحالة', 'التاريخ'],
        'data_fn': lambda: Purchase.query.order_by(Purchase.created_at.desc()).all(),
        'row_fn': lambda r: [r.id, r.invoice_number, r.supplier.name if r.supplier else '-', r.store.name if r.store else '-', r.total, r.paid, r.remaining, r.status, r.created_at.strftime('%Y-%m-%d')],
    },
    'sales': {
        'title': 'المبيعات',
        'columns': ['#', 'رقم الفاتورة', 'العميل', 'المخزن', 'الإجمالي', 'الخصم', 'المدفوع', 'المتبقي', 'الحالة', 'التاريخ'],
        'data_fn': lambda: Sale.query.order_by(Sale.created_at.desc()).all(),
        'row_fn': lambda r: [r.id, r.invoice_number, r.customer.name if r.customer else 'نقدي', r.store.name if r.store else '-', r.total, r.discount, r.paid, r.remaining, r.status, r.created_at.strftime('%Y-%m-%d')],
    },
    'returns': {
        'title': 'المرتجعات',
        'columns': ['#', 'رقم المرتجع', 'العميل', 'فاتورة المبيعات', 'الإجمالي', 'السبب', 'التاريخ'],
        'data_fn': lambda: Return.query.order_by(Return.created_at.desc()).all(),
        'row_fn': lambda r: [r.id, r.return_number, r.customer.name if r.customer else '-', r.sale.invoice_number if r.sale else '-', r.total, r.reason or '-', r.created_at.strftime('%Y-%m-%d')],
    },
    'expenses': {
        'title': 'المصروفات',
        'columns': ['#', 'الوصف', 'المبلغ', 'التصنيف', 'الفرع', 'ملاحظات', 'التاريخ'],
        'data_fn': lambda: Expense.query.order_by(Expense.created_at.desc()).all(),
        'row_fn': lambda r: [r.id, r.description, r.amount, r.category or '-', r.branch.name if r.branch else '-', r.note or '-', r.created_at.strftime('%Y-%m-%d')],
    },
    'bonds': {
        'title': 'السندات',
        'columns': ['#', 'رقم السند', 'النوع', 'المبلغ', 'العميل', 'المورد', 'ملاحظات', 'التاريخ'],
        'data_fn': lambda: Bond.query.order_by(Bond.created_at.desc()).all(),
        'row_fn': lambda r: [r.id, r.bond_number, 'قبض' if r.bond_type == 'receipt' else 'صرف', r.amount, r.customer.name if r.customer else '-', r.supplier.name if r.supplier else '-', r.note or '-', r.created_at.strftime('%Y-%m-%d')],
    },
    'employees': {
        'title': 'الموظفين',
        'columns': ['#', 'الاسم', 'البريد', 'الهاتف', 'الفرع', 'عدد الصلاحيات', 'الحالة'],
        'data_fn': lambda: Employee.query.order_by(Employee.created_at.desc()).all(),
        'row_fn': lambda r: [r.id, r.user.full_name if r.user else '-', r.user.email if r.user else '-', r.phone or '-', r.branch.name if r.branch else 'جميع الفروع', len([v for v in r.get_permissions().values() if v]), 'نشط' if r.is_active else 'غير نشط'],
    },
}


def generate_excel(export_type):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    config = EXPORT_CONFIG.get(export_type)
    if not config:
        return None

    data = config['data_fn']()
    wb = Workbook()
    ws = wb.active
    ws.title = config['title']

    # Colors
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_font = Font(name='Cairo', size=12, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')
    cell_font = Font(name='Cairo', size=11)
    cell_align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )
    alt_fill = PatternFill(start_color='F8F9FC', end_color='F8F9FC', fill_type='solid')

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(config['columns']))
    title_cell = ws.cell(row=1, column=1, value=config['title'])
    title_cell.font = Font(name='Cairo', size=16, bold=True, color='4F46E5')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(config['columns']))
    sub_cell = ws.cell(row=2, column=1, value=f'Mazad Plus - {datetime.utcnow().strftime("%Y-%m-%d %H:%M")}')
    sub_cell.font = Font(name='Cairo', size=10, color='9CA3AF')
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25

    # Header row
    header_row = 4
    for col_idx, col_name in enumerate(config['columns'], 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[header_row].height = 35

    # Data rows
    for row_idx, record in enumerate(data, header_row + 1):
        row_data = config['row_fn'](record)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = border
            if (row_idx - header_row) % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 28

    # Auto-width columns
    for col_idx in range(1, len(config['columns']) + 1):
        max_len = len(config['columns'][col_idx - 1])
        for row in ws.iter_rows(min_row=header_row, min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_val in row:
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    # Freeze panes
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf(export_type):
    from fpdf import FPDF
    import arabic_reshaper
    from bidi.algorithm import get_display

    config = EXPORT_CONFIG.get(export_type)
    if not config:
        return None

    data = config['data_fn']()

    FONT_REG = 'C:/Windows/Fonts/arial.ttf'
    FONT_BOLD = 'C:/Windows/Fonts/arialbd.ttf'
    FONT_ITALIC = 'C:/Windows/Fonts/ariali.ttf'

    def ar(text):
        if text is None:
            text = ''
        try:
            reshaped = arabic_reshaper.reshape(str(text))
            return get_display(reshaped)
        except:
            return str(text)

    class ExportPDF(FPDF):
        def header(self):
            self.set_font('Arial', 'B', 16)
            self.set_text_color(79, 70, 229)
            self.cell(0, 10, ar(config['title']), align='C', new_x='LMARGIN', new_y='NEXT')
            self.set_font('Arial', '', 9)
            self.set_text_color(156, 163, 175)
            self.cell(0, 6, ar(f'Mazad Plus - {datetime.utcnow().strftime("%Y-%m-%d %H:%M")}'), align='C', new_x='LMARGIN', new_y='NEXT')
            self.line(10, self.get_y(), 290, self.get_y())
            self.ln(4)

        def footer(self):
            self.set_y(-15)
            self.set_font('Arial', '', 8)
            self.set_text_color(156, 163, 175)
            self.cell(0, 10, ar(f'إجمالي السجلات: {len(data)} | تم التصدير بواسطة: {current_user.full_name} | الصفحة {self.page_no()}/{{nb}}'), align='C')

    pdf = ExportPDF(orientation='L', unit='mm', format='A4')
    pdf.add_font('Arial', '', FONT_REG, uni=True)
    pdf.add_font('Arial', 'B', FONT_BOLD, uni=True)
    pdf.add_font('Arial', 'I', FONT_ITALIC, uni=True)
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    col_count = len(config['columns'])
    col_width = 270 / col_count

    pdf.set_font('Arial', 'B', 9)
    pdf.set_fill_color(79, 70, 229)
    pdf.set_text_color(255, 255, 255)
    for col_name in config['columns']:
        pdf.cell(col_width, 8, ar(col_name), border=1, fill=True, align='C')
    pdf.ln()

    pdf.set_font('Arial', '', 8)
    for idx, record in enumerate(data):
        row_data = config['row_fn'](record)
        if idx % 2 == 0:
            pdf.set_fill_color(248, 249, 252)
        else:
            pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(31, 41, 55)

        for val in row_data:
            if isinstance(val, float):
                formatted = '{:,.2f}'.format(val)
            else:
                formatted = str(val)
            pdf.cell(col_width, 7, ar(formatted), border=1, fill=True, align='C')
        pdf.ln()

    import io
    return io.BytesIO(pdf.output())


@app.route('/export/<export_type>/<fmt>')
@login_required
def export_data(export_type, fmt):
    if export_type not in EXPORT_CONFIG:
        flash('نوع التصدير غير صالح', 'danger')
        return redirect(url_for('dashboard'))

    if fmt == 'excel':
        output = generate_excel(export_type)
        if not output:
            flash('فشل إنشاء ملف Excel', 'danger')
            return redirect(url_for(export_type + '_list' if export_type != 'employees' else export_type))
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{export_type}_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.xlsx',
        )
    elif fmt == 'pdf':
        pdf = generate_pdf(export_type)
        if not pdf:
            flash('فشل إنشاء ملف PDF', 'danger')
            return redirect(url_for(export_type + '_list' if export_type != 'employees' else export_type))
        from flask import send_file
        return send_file(
            pdf,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'{export_type}_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.pdf',
        )
    else:
        flash('صيغة التصدير غير مدعومة', 'danger')
        return redirect(url_for('dashboard'))


# ═══════════════════════════════════════════════
# API Routes
# ═══════════════════════════════════════════════

@app.route('/api/check-auth')
@csrf.exempt
def check_auth():
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        user = verify_token(auth[7:])
        if user:
            login_user(user)
    if current_user.is_authenticated:
        return jsonify({
            'authenticated': True,
            'user': {
                'id': current_user.id,
                'name': current_user.full_name,
                'email': current_user.email,
                'role': current_user.role
            }
        })
    return jsonify({'authenticated': False}), 401


@app.route('/api/items-by-store/<int:store_id>')
@csrf.exempt
@login_required
def api_items_by_store(store_id):
    store = Store.query.filter_by(id=store_id).first()
    if not store or (store.user_id and store.user_id != current_user.id and not current_user.is_super_admin):
        return jsonify([])
    items = Item.query.filter_by(store_id=store_id, is_active=True, user_id=current_user.id).all()
    return jsonify([{
        'id': i.id,
        'name': i.name,
        'barcode': i.barcode,
        'sell_price': i.sell_price,
        'buy_price': i.buy_price,
        'quantity': i.quantity,
    } for i in items])


@app.route('/api/invoice-items/<string:type>/<int:id>')
@csrf.exempt
@login_required
def api_invoice_items(type, id):
    if type == 'purchase':
        obj = Purchase.query.get_or_404(id)
        if obj.user_id and obj.user_id != current_user.id and not current_user.is_super_admin:
            return jsonify({'error': 'غير مصرح'}), 403
        items_data = [{
            'item_id': pi.item_id,
            'item_name': pi.item.name if pi.item else '',
            'quantity': pi.quantity,
            'price': pi.price,
            'total': pi.total,
        } for pi in obj.items]
        return jsonify({'items': items_data, 'total': obj.total, 'paid': obj.paid, 'note': obj.note})
    elif type == 'sale':
        obj = Sale.query.get_or_404(id)
        if obj.user_id and obj.user_id != current_user.id and not current_user.is_super_admin:
            return jsonify({'error': 'غير مصرح'}), 403
        items_data = [{
            'item_id': si.item_id,
            'item_name': si.item.name if si.item else '',
            'quantity': si.quantity,
            'price': si.price,
            'total': si.total,
        } for si in obj.items]
        return jsonify({'items': items_data, 'total': obj.total, 'paid': obj.paid, 'discount': obj.discount, 'note': obj.note})
    return jsonify({'error': 'Invalid type'}), 400


# ═══════════════════════════════════════════════
# Notification System (Admin send only)
# ═══════════════════════════════════════════════


@app.route('/admin/notifications')
@admin_required
def admin_notifications():
    users = User.query.filter(User.is_super_admin == False).order_by(User.full_name).all()
    sent = Notification.query.order_by(Notification.created_at.desc()).limit(100).all()
    return render_template('admin_notifications.html', users=users, sent=sent)


@app.route('/admin/notifications/send', methods=['POST'])
@admin_required
def admin_send_notification():
    recipient_id = request.form.get('recipient_id')
    title = request.form.get('title', '').strip()
    message = request.form.get('message', '').strip()
    notif_type = request.form.get('type', 'info')
    link = request.form.get('link', '').strip() or None

    if not title:
        flash('عنوان الإشعار مطلوب', 'danger')
        return redirect(url_for('admin_notifications'))

    if recipient_id == 'all':
        targets = User.query.filter(User.is_super_admin == False, User.is_active == True).all()
        for user in targets:
            db.session.add(Notification(
                sender_id=current_user.id,
                recipient_id=user.id,
                title=title,
                message=message,
                type=notif_type,
                link=link,
            ))
        flash(f'تم إرسال الإشعار إلى {len(targets)} مستخدم', 'success')
    else:
        user = db.session.get(User, int(recipient_id))
        if not user:
            flash('المستخدم غير موجود', 'danger')
            return redirect(url_for('admin_notifications'))
        db.session.add(Notification(
            sender_id=current_user.id,
            recipient_id=user.id,
            title=title,
            message=message,
            type=notif_type,
            link=link,
        ))
        flash(f'تم إرسال الإشعار إلى {user.full_name}', 'success')

    safe_commit()
    return redirect(url_for('admin_notifications'))


def time_ago(dt):
    diff = datetime.utcnow() - dt
    seconds = int(diff.total_seconds())
    if seconds < 60:
        return 'الآن'
    minutes = seconds // 60
    if minutes < 60:
        return f'منذ {minutes} دقيقة'
    hours = minutes // 60
    if hours < 24:
        return f'منذ {hours} ساعة'
    days = hours // 24
    if days < 30:
        return f'منذ {days} يوم'
    return dt.strftime('%Y-%m-%d')


# ─────────────────────────── Init DB ───────────────────────────

def init_database():
    with app.app_context():
        db.create_all()

        # ── Migrate: add trial_days column to plans if missing ──
        try:
            from sqlalchemy import text
            result = db.session.execute(text('PRAGMA table_info(plans)')).fetchall()
            col_names = [row[1] for row in result]
            if 'trial_days' not in col_names:
                db.session.execute(text('ALTER TABLE plans ADD COLUMN trial_days INTEGER DEFAULT 7'))
                db.session.commit()
                print('✓ Added trial_days column to plans table')
        except Exception as e:
            db.session.rollback()
            print(f'⚠ Migration note: {e}')

        # ── Migrate: add barcode column to purchases if missing ──
        try:
            result = db.session.execute(text('PRAGMA table_info(purchases)')).fetchall()
            col_names = [row[1] for row in result]
            if 'barcode' not in col_names:
                db.session.execute(text('ALTER TABLE purchases ADD COLUMN barcode VARCHAR(50)'))
                db.session.commit()
        except Exception:
            db.session.rollback()

        # ── Migrate: add barcode column to sales if missing ──
        try:
            result = db.session.execute(text('PRAGMA table_info(sales)')).fetchall()
            col_names = [row[1] for row in result]
            if 'barcode' not in col_names:
                db.session.execute(text('ALTER TABLE sales ADD COLUMN barcode VARCHAR(50)'))
                db.session.commit()
        except Exception:
            db.session.rollback()

        # ── Migrate: add store_logo column to store_settings if missing ──
        try:
            result = db.session.execute(text('PRAGMA table_info(store_settings)')).fetchall()
            col_names = [row[1] for row in result]
            if 'store_logo' not in col_names:
                db.session.execute(text('ALTER TABLE store_settings ADD COLUMN store_logo VARCHAR(500)'))
                db.session.commit()
                print('✓ Added store_logo column to store_settings table')
        except Exception:
            db.session.rollback()

        # ── Migrate: add user_id column to branches, stores, categories, items, customers, suppliers ──
        _migrate_cols = {
            'branches': 'user_id',
            'stores': 'user_id',
            'categories': 'user_id',
            'items': 'user_id',
            'customers': 'user_id',
            'suppliers': 'user_id',
        }
        for table, col in _migrate_cols.items():
            try:
                result = db.session.execute(text(f'PRAGMA table_info({table})')).fetchall()
                col_names = [row[1] for row in result]
                if col not in col_names:
                    db.session.execute(text(f'ALTER TABLE {table} ADD COLUMN {col} INTEGER REFERENCES users(id)'))
                    db.session.commit()
                    print(f'✓ Added {col} column to {table} table')
            except Exception:
                db.session.rollback()

        # ── Migrate: add email_verified, google_id, verification_token to users ──
        try:
            result = db.session.execute(text('PRAGMA table_info(users)')).fetchall()
            col_names = [row[1] for row in result]
            for col in ('email_verified', 'google_id', 'verification_token'):
                if col not in col_names:
                    db.session.execute(text(f'ALTER TABLE users ADD COLUMN {col} VARCHAR(100)'))
                    db.session.commit()
        except Exception:
            db.session.rollback()

        # Use raw SQL for initial queries to avoid ORM schema mismatch
        from sqlalchemy import text as sqltext
        user_count = db.session.execute(sqltext('SELECT COUNT(*) FROM users')).scalar()
        if user_count == 0:
            # Create admin using raw SQL with hashed password
            from werkzeug.security import generate_password_hash
            pw_hash = generate_password_hash('admin123')
            db.session.execute(
                sqltext('INSERT INTO users (full_name, email, phone, password_hash, role, is_super_admin, is_active, created_at) VALUES (:fn, :em, :ph, :pw, :rl, 1, 1, datetime(\'now\'))'),
                {'fn': 'مدير النظام', 'em': 'admin@mazadplus.com', 'ph': '01000000000', 'pw': pw_hash, 'rl': 'admin'}
            )
            db.session.commit()
            print('✓ Super Admin: admin@mazadplus.com / admin123')

        plan_count = db.session.execute(sqltext('SELECT COUNT(*) FROM plans')).scalar()
        if plan_count == 0:
            plans_data = [
                ('الخطة الأساسية', 'للمتاجر الصغيرة', 0, 30, 1, 1, 50, 50, 20, 100, 1, 'مخزن واحد\n50 صنف كحد أقصى\n100 فاتورة شهرياً\nدعم فني عبر البريد', 7),
                ('الخطة المهنية', 'للمتاجر المتوسطة', 299, 30, 3, 3, 500, 200, 100, 500, 3, '3 مخازن\n500 صنف\n500 فاتورة\n3 مستخدمين\nدعم فني', 7),
                ('الخطة المتقدمة', 'للمتاجر الكبيرة', 799, 30, 10, 10, 2000, 1000, 500, 2000, 10, '10 مخازن\n2000 صنف\n2000 فاتورة\nتقارير وتحليلات', 7),
                ('الخطة الغير محدودة', 'بدون قيود', 1599, 30, 999, 999, 99999, 99999, 99999, 99999, 50, 'غير محدود\nكل المميزات مفتوحة', 7),
            ]
            for p in plans_data:
                db.session.execute(
                    sqltext('INSERT INTO plans (name, description, price, duration_days, max_branches, max_stores, max_items, max_customers, max_suppliers, max_invoices_monthly, max_users, features, trial_days, is_active, created_at) VALUES (:n, :d, :pr, :dd, :mb, :ms, :mi, :mc, :msup, :minv, :mu, :f, :td, 1, datetime(\'now\'))'),
                    {'n': p[0], 'd': p[1], 'pr': p[2], 'dd': p[3], 'mb': p[4], 'ms': p[5], 'mi': p[6], 'mc': p[7], 'msup': p[8], 'minv': p[9], 'mu': p[10], 'f': p[11], 'td': p[12]}
                )
            db.session.commit()
            print('✓ 4 default plans created')


if __name__ == '__main__':
    init_database()
    debug_mode = os.environ.get('FLASK_DEBUG', '0') == '1'
    app.run(debug=debug_mode, host='0.0.0.0', port=5000)
